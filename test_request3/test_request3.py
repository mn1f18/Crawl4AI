import openpyxl
import asyncio
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig, BrowserConfig, CacheMode
from bs4 import BeautifulSoup
import os
import json
import re
import traceback
from datetime import datetime
from urllib.parse import urlparse, urljoin, urlunparse
from crawl4ai.extraction_strategy import JsonCssExtractionStrategy
from crawl4ai import CacheMode
import requests
import pandas as pd
import time
import hashlib
from concurrent.futures import ThreadPoolExecutor
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
from crawl4ai.content_filter_strategy import PruningContentFilter
import random

# 导入AI链接验证模块
try:
    # 尝试从同级目录导入
    from ai_link_validator import is_valid_news_link_with_ai, batch_link_validation
    AI_LINK_VALIDATOR_AVAILABLE = True
except ImportError:
    try:
        # 尝试从上级目录导入
        import sys
        import os
        sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from ai_link_validator import is_valid_news_link_with_ai, batch_link_validation
        AI_LINK_VALIDATOR_AVAILABLE = True
    except ImportError:
        print("⚠️ AI链接验证模块未找到，将使用传统方法判断链接有效性")
        AI_LINK_VALIDATOR_AVAILABLE = False

# 配置选项
SAVE_TO_EXCEL = True                # 是否保存结果到Excel
MAX_LINKS_PER_SOURCE = 20           # 每个来源最多抓取的链接数
MIN_CONTENT_LENGTH = 300            # 最小有效内容长度
MAX_RETRY_COUNT = 2                 # 链接请求失败时最大重试次数
REQUEST_TIMEOUT = 60                # 请求超时时间（秒）
SKIP_EXISTING_LINKS = True          # 是否跳过已存在的链接
LINKS_HISTORY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "link_history")  # 历史链接存储目录
USE_AI_LINK_VALIDATION = True       # 是否使用AI链接验证
USE_COLD_START = False               # 冷启动模式：只收集链接到历史记录，不进行内容爬取（适用于首次遇到大量链接时）

# 简化评分参数 - 仅保留基础设置
QUALITY_CONFIG = {
    "min_content_length": 300,    # 最小有效内容长度
    "min_paragraphs": 3           # 最少段落数
}

# 确保历史目录存在
if not os.path.exists(LINKS_HISTORY_DIR):
    os.makedirs(LINKS_HISTORY_DIR)

# 使用URL作为链接唯一标识
def generate_link_id(url):
    """生成链接的唯一标识，直接使用URL作为标识"""
    return url

# 添加URL规范化函数
def normalize_url(url):
    """规范化URL以便进行比较，移除末尾斜杠和www.前缀等，优化高频调用性能"""
    if not url:
        return ""
    
    try:
        # 快速检测是否为URL，避免不必要的解析
        if not any(p in url for p in ('http://', 'https://', 'www.')):
            return url.strip().rstrip('/')
        
        # 解析URL
        parsed = urlparse(url)
        
        # 获取域名和路径部分
        netloc = parsed.netloc
        path = parsed.path
        
        # 移除www.前缀
        if netloc.startswith('www.'):
            netloc = netloc[4:]
        
        # 移除末尾斜杠
        if path.endswith('/') and len(path) > 1:
            path = path[:-1]
        
        # 重新组合URL
        # 优化: 只保留scheme, netloc, path部分，忽略query和fragment
        normalized = f"{parsed.scheme}://{netloc}{path}"
        
        return normalized
    except Exception:
        # 出错时返回原始URL
        return url.strip()

# 检测是否为新链接并追踪链接状态
def is_new_link(url, source):
    """
    检查链接是否为新链接（未爬取过）并获取状态
    
    返回: (is_new, is_invalid, is_processed)
    is_new: 布尔值，True表示链接未爬取过
    is_invalid: 布尔值，True表示链接之前已被标记为无效
    is_processed: 布尔值，True表示链接已经被处理过（无论有效或无效）
    """
    links_history = load_links_history(source)
    
    # 直接检查URL是否存在于历史记录中
    for _, info in links_history.items():
        if info.get('url', '') == url:
            is_valid = info.get('is_valid', True)
            # 链接存在且已经抓取过内容（内容长度大于0或爬取次数大于1）
            has_content = info.get('content_length', 0) > 0 or info.get('crawl_count', 0) > 1
            return False, not is_valid, has_content
    
    # URL不存在于历史记录中，则为新链接且未处理过
    return True, False, False

# 加载历史链接数据
def load_links_history(source):
    """加载指定来源的历史链接数据，使用更高效的方式"""
    history_file = os.path.join(LINKS_HISTORY_DIR, f"{source}_links.json")
    links_data = {}
    
    try:
        if os.path.exists(history_file):
            # 检查文件大小以决定加载方式
            file_size = os.path.getsize(history_file)
            
            with open(history_file, 'r', encoding='utf-8') as f:
                # 对于大文件，使用更高效的方式读取
                if file_size > 10 * 1024 * 1024:  # 大于10MB的文件
                    print(f"⚠️ 历史链接文件较大 ({file_size/1024/1024:.2f}MB)，使用优化加载方式...")
                    links_data = json.load(f)
                else:
                    # 小文件则正常加载
                    links_data = json.load(f)
                    
            print(f"✅ 已加载 {len(links_data)} 条历史链接: {history_file}")
        else:
            print(f"⚠️ 未找到历史链接文件，将创建新的记录: {history_file}")
    except Exception as e:
        print(f"❌ 加载历史链接数据失败: {str(e)}")
        traceback.print_exc()
    
    return links_data

# 保存历史链接数据
def save_links_history(source, links_data):
    """保存链接历史记录到本地JSON文件，使用更高效的批量保存方式"""
    if not source or not links_data:
        return False
        
    # 确保目录存在
    try:
        os.makedirs(LINKS_HISTORY_DIR, exist_ok=True)
        
        # 构建文件路径
        file_path = os.path.join(LINKS_HISTORY_DIR, f"{source}_links.json")
        
        # 使用临时文件进行写入，然后重命名，避免文件损坏
        temp_file = file_path + ".tmp"
        
        # 优化: 如果数据不超过一定规模，直接使用标准JSON保存
        if len(links_data) < 10000:  # 对于小规模数据
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(links_data, f, ensure_ascii=False, indent=2)
        else:
            # 对于大规模数据，使用更高效的方式
            with open(temp_file, 'w', encoding='utf-8') as f:
                # 不使用缩进，减少文件大小和保存时间
                json.dump(links_data, f, ensure_ascii=False, separators=(',', ':'))
        
        # 安全地替换原文件
        if os.path.exists(temp_file):
            # 在Windows系统中，如果目标文件存在需要先删除
            if os.path.exists(file_path):
                os.remove(file_path)
            os.rename(temp_file, file_path)
            
        return True
    except Exception as e:
        print(f"❌ 保存链接历史记录失败: {str(e)}")
        traceback.print_exc()
        return False

# 更新链接历史记录
def update_link_history(url, title, source, content_summary="", is_valid=True, quality_score=0, content_length=0, error_message="", content_fingerprint="", ai_score=0, ai_reason="", link_type=""):
    """
    更新链接历史记录
    
    参数:
        url: 链接URL
        title: 链接标题
        source: 来源ID
        content_summary: 内容摘要
        is_valid: 是否有效的链接
        quality_score: 内容质量评分
        content_length: 内容长度
        error_message: 错误信息
        content_fingerprint: 内容指纹
        ai_score: AI验证分数
        ai_reason: AI验证理由
        link_type: 链接类型
    """
    try:
        # 加载历史记录
        links_history = load_links_history(source)
        
        # 生成链接ID
        link_id = generate_link_id(url)
        
        # 当前时间
        current_time = datetime.now().isoformat()
        
        # 检查此链接是否存在
        if link_id in links_history:
            # 更新现有记录
            links_history[link_id]['is_valid'] = is_valid
            links_history[link_id]['last_updated'] = current_time
            
            # 只有在提供了内容时才更新内容相关字段
            if content_length > 0:
                links_history[link_id]['content_length'] = content_length
                if content_summary:
                    links_history[link_id]['content_summary'] = content_summary
                if content_fingerprint:
                    links_history[link_id]['content_fingerprint'] = content_fingerprint
            
            # 更新质量评分（如果提供）
            if quality_score > 0:
                links_history[link_id]['quality_score'] = quality_score
                
            # 更新错误信息（如果有）
            if error_message:
                links_history[link_id]['error_message'] = error_message
                
            # 增加爬取次数
            links_history[link_id]['crawl_count'] = links_history[link_id].get('crawl_count', 0) + 1
                
            # 更新AI验证信息（如果提供）
            if ai_score > 0:
                links_history[link_id]['ai_score'] = ai_score
            if ai_reason:
                links_history[link_id]['ai_reason'] = ai_reason
            if link_type:
                links_history[link_id]['link_type'] = link_type
        else:
            # 创建新记录
            links_history[link_id] = {
                'url': url,
                'title': title,
                'is_valid': is_valid,
                'first_seen': current_time,
                'last_updated': current_time,
                'content_length': content_length,
                'content_summary': content_summary,
                'quality_score': quality_score,
                'error_message': error_message,
                'content_fingerprint': content_fingerprint,
                'crawl_count': 1,  # 初始爬取次数为1
                'ai_score': ai_score,
                'ai_reason': ai_reason,
                'link_type': link_type
            }
            
        # 保存更新后的历史记录
        save_links_history(source, links_history)
        
    except Exception as e:
        print(f"更新链接历史记录出错: {e}")

# 读取主界面链接
# 使用当前脚本目录下的testhomepage.xlsx文件
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_source_file = os.path.join(current_dir, "testhomepage.xlsx")
# 备选相对路径，如果当前目录下不存在可以尝试上级目录
alternative_path = '../testhomepage.xlsx'

# 检查文件是否存在，如果不存在尝试备选路径
if not os.path.exists(excel_source_file):
    print(f"❌ 主路径找不到文件: {excel_source_file}，尝试备选路径...")
    if os.path.exists(alternative_path):
        excel_source_file = alternative_path
        print(f"✅ 找到备选路径文件: {excel_source_file}")
    else:
        print(f"❌ 备选路径也找不到文件: {alternative_path}")
        # 尝试列出当前目录下的Excel文件
        print("当前目录下的Excel文件:")
        for file in os.listdir('.'):
            if file.endswith('.xlsx'):
                print(f" - {file}")
        
        # 尝试列出上级目录下的Excel文件
        print("上级目录下的Excel文件:")
        for file in os.listdir('..'):
            if file.endswith('.xlsx'):
                print(f" - {file}")
else:
    print(f"✅ 找到源文件: {excel_source_file}")

wb = openpyxl.load_workbook(excel_source_file)
sheet = wb.active

# 创建结果Excel文件
def create_result_excel(filename=None):
    """创建结果Excel文件"""
    # 如果不保存到Excel，直接返回None
    if not SAVE_TO_EXCEL:
        return None
    
    # 生成文件名，确保使用当前目录路径
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"news_results_{timestamp}.xlsx"
    
    # 确保文件保存在正确的目录下
    if not os.path.isabs(filename):
        # 使用当前脚本所在目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        filename = os.path.join(current_dir, filename)
    
    # 保存全局引用
    global result_excel_file
    result_excel_file = filename
    
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "抓取结果"
    
    # 设置表头
    headers = [
        '索引', '来源', '链接', '标题', '发布日期', '分类', '正文字数', 
        '爬取时间(秒)', '状态', '尝试次数', 'AI验证分数', 'AI验证理由', '链接类型'
    ]
    
    # 设置列宽
    column_widths = {
        'A': 8,   # 索引
        'B': 15,  # 来源
        'C': 50,  # 链接
        'D': 50,  # 标题
        'E': 15,  # 发布日期
        'F': 15,  # 分类
        'G': 10,  # 正文字数
        'H': 15,  # 爬取时间
        'I': 20,  # 状态
        'J': 10,  # 尝试次数
        'K': 12,  # AI验证分数
        'L': 40,  # AI验证理由
        'M': 15,  # 链接类型
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # 写入表头
    for idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=idx).value = header
    
    # 保存并返回
    wb.save(filename)
    print(f"已创建结果文件: {filename}")
    return filename

# 生成当前批次ID
current_batch = datetime.now().strftime("%Y%m%d_%H%M%S")
batch_id = current_batch

# 全局变量 - 修改默认路径使用当前目录
current_dir = os.path.dirname(os.path.abspath(__file__))
result_excel_file = os.path.join(current_dir, f"news_results_{batch_id}.xlsx")  # 默认Excel文件名使用完整路径
LOG_FILE = os.path.join(current_dir, "ai_link_decisions.jsonl")  # AI验证日志文件路径
ENABLE_LOGGING = True  # 是否启用AI验证日志

# 简化评估内容质量的函数
def evaluate_content_quality(html_content, title, url=""):
    """
    简化版函数，不再评估内容质量，只返回内容摘要和指纹
    """
    if not html_content:
        return True, "", ""
    
    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # 移除脚本、样式和导航元素
    for tag in soup(['script', 'style', 'nav', 'header', 'footer', 'aside']):
        tag.decompose()
    
    # 获取所有文本
    text = soup.get_text(separator='\n', strip=True)
    
    # 生成内容摘要 (限制为200字)
    content_summary = text[:200] + "..." if len(text) > 200 else text
    
    # 生成内容指纹
    content_fingerprint = hashlib.md5(text[:1000].encode('utf-8')).hexdigest()
    
    # 始终返回True表示内容有效
    return True, content_summary, content_fingerprint

# 提取发布时间的函数（增强版）
def extract_publish_date(soup):
    """
    从HTML中提取发布时间，尝试多种方法
    """
    # 方法1: 查找元数据标签
    meta_properties = [
        'article:published_time', 'datePublished', 'publishedDate', 
        'pubdate', 'date', 'DC.date', 'article:modified_time', 
        'lastModified', 'og:published_time', 'og:updated_time',
        'dateCreated', 'dateModified', 'release_date'
    ]
    
    for meta in soup.find_all('meta'):
        prop = meta.get('property', '') or meta.get('name', '')
        if prop.lower() in [p.lower() for p in meta_properties]:
            content = meta.get('content')
            if content:
                return content
    
    # 方法2: 查找时间标签
    time_tags = soup.find_all('time')
    for time_tag in time_tags:
        datetime_attr = time_tag.get('datetime')
        if datetime_attr:
            return datetime_attr
        if time_tag.text.strip():
            return time_tag.text.strip()
    
    # 方法3: 根据CSS选择器查找
    date_selectors = [
        '.date', '.time', '.published', '.timestamp', '.post-date',
        '[itemprop="datePublished"]', '.article-date', '.entry-date',
        '.post-meta time', '.article-meta time', '.publish-date',
        '.updated-date', '.create-date', '.release-date', '.article-time'
    ]
    
    for selector in date_selectors:
        date_element = soup.select_one(selector)
        if date_element:
            return re.sub(r'(Published|Updated|Posted|Date):?\s*', '', date_element.get_text(strip=True))
    
    # 方法4: 在全文中查找日期模式
    # 常见日期格式：YYYY-MM-DD, DD/MM/YYYY, Month DD, YYYY等
    date_patterns = [
        r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
        r'\d{2}/\d{2}/\d{4}',  # DD/MM/YYYY
        r'\d{2}\.\d{2}\.\d{4}', # DD.MM.YYYY
        r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}',
        r'\d{1,2}\s+(January|February|March|April|May|June|July|August|September|October|November|December),?\s+\d{4}'
    ]
    
    text = soup.get_text()
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0)
    
    # 没有找到日期
    return "未知"

# 提取文章标题的函数（增强版）
def extract_title(soup, url):
    """
    从HTML中提取文章标题，尝试多种方法
    """
    # 方法1: 使用<title>标签
    if soup.title:
        title = soup.title.string
        # 清理标题（移除网站名称等）
        if title:
            # 常见分隔符：|, -, –, —, :, ·, •
            for separator in [' | ', ' - ', ' – ', ' — ', ' : ', ' · ', ' • ']:
                if separator in title:
                    parts = title.split(separator)
                    # 通常第一部分是文章标题，而非网站名称
                    if len(parts[0]) > 10:  # 标题足够长
                        return parts[0].strip()
            return title.strip()
    
    # 方法2: 寻找主标题标签
    for heading in ['h1', 'h2']:
        headings = soup.find_all(heading)
        if headings:
            # 筛选最可能是标题的元素（通常位于内容区域，不在header/nav等中）
            main_content = soup.find('main') or soup.find('article') or soup.find('div', class_=lambda c: c and ('content' in c.lower() or 'article' in c.lower()))
            
            if main_content:
                content_headings = main_content.find_all(heading)
                if content_headings:
                    return content_headings[0].get_text(strip=True)
            
            # 如果不能确定内容区域，使用第一个最长的标题
            longest_heading = max(headings, key=lambda h: len(h.get_text(strip=True)))
            return longest_heading.get_text(strip=True)
    
    # 方法3: 查找元数据
    meta_tags = soup.find_all('meta')
    for meta in meta_tags:
        prop = meta.get('property', '') or meta.get('name', '')
        if prop.lower() in ['og:title', 'twitter:title', 'dc.title']:
            content = meta.get('content')
            if content:
                return content
    
    # 方法4: 从URL中提取可能的标题
    path = urlparse(url).path
    segments = [seg for seg in path.split('/') if seg]
    if segments:
        # 最后一段路径可能是标题的slug
        last_segment = segments[-1]
        # 将连字符或下划线替换为空格
        title_from_url = re.sub(r'[-_]', ' ', last_segment)
        if len(title_from_url) > 5:  # 确保足够长
            return title_from_url.capitalize()
    
    # 无法确定标题
    return "未找到标题"

# 判断链接是否为有效新闻链接的函数（简化版）
def is_valid_news_link(link, base_url, a_tag=None, html_content=None):
    """
    判断链接是否为有效的新闻链接 - 简化版，主要依赖AI
    
    参数:
        link: 待验证的链接
        base_url: 基础链接
        a_tag: 可选，链接所在的a标签（BeautifulSoup对象）
        html_content: 可选，页面HTML内容
    """
    # 在冷启动模式下，所有链接都视为有效
    if USE_COLD_START:
        return True
        
    # 基础过滤 - 明显的非新闻链接
    if not link or not isinstance(link, str):
        return False
        
    if link.startswith(('#', 'javascript:', 'mailto:', 'tel:')):
        return False
        
    # 构建完整URL
    if not urlparse(link).netloc:
        full_url = urljoin(base_url, link)
        parsed_link = urlparse(full_url)
    else:
        full_url = link
        parsed_link = urlparse(full_url)
        
    # 排除非文本媒体文件
    if any(parsed_link.path.lower().endswith(ext) for ext in 
          ['.jpg', '.jpeg', '.png', '.gif', '.pdf', '.zip', '.mp3', '.mp4', 
           '.css', '.js', '.ico', '.svg', '.webp', '.woff', '.ttf']):
        return False
    
    # 检查是否为同一域名（排除外部链接）
    parsed_base = urlparse(base_url)
    if parsed_link.netloc != parsed_base.netloc:
        return False

    # 优先使用AI验证模块
    if USE_AI_LINK_VALIDATION and AI_LINK_VALIDATOR_AVAILABLE and not USE_COLD_START:
        try:
            # 使用AI验证模块判断链接有效性
            from ai_link_validator import is_valid_news_link_with_ai
            return is_valid_news_link_with_ai(full_url, base_url, a_tag, html_content)
        except Exception as e:
            print(f"❌ AI链接验证失败，回退到基础规则: {str(e)}")
    
    # 基础过滤规则（如果AI验证不可用或失败）
    # 检查链接文本是否看起来像标题（超过20个字符且不是导航类文本）
    if a_tag and a_tag.get_text():
        link_text = a_tag.get_text().strip()
        if len(link_text) > 15:  # 放宽条件
            return True
    
    # 检查URL路径是否包含新闻相关特征
    path_lower = parsed_link.path.lower()
    return (
        '/news/' in path_lower or 
        '/article/' in path_lower or 
        '/post/' in path_lower or
        re.search(r'/20\d{2}[/-]\d{1,2}/', path_lower) is not None or
        re.search(r'/\d{4,}/', path_lower) is not None
    )

# 爬取主界面
async def fetch_news_links(main_url, source):
    """获取新闻链接，优化链接比对过程"""
    # 初始化统计
    link_counts = {
        "found_links": 0,
        "new_links": 0,
        "valid_links": 0,
        "invalid_links": 0,
        "skipped_existing": 0,
        "skipped_invalid": 0,
        "skipped_processed": 0,
        "total_ai_validated": 0
    }
    
    try:
        print(f"🔄 正在爬取主页: {main_url}")
        
        # 设置爬虫配置
        prune_filter = PruningContentFilter(
            threshold=0.45,
            threshold_type="dynamic",
            min_word_threshold=5
        )
        md_generator = DefaultMarkdownGenerator(content_filter=prune_filter)
        
        crawler_config = CrawlerRunConfig(
            markdown_generator=md_generator,
            page_timeout=REQUEST_TIMEOUT * 1000,
            cache_mode=CacheMode.BYPASS  # 禁用缓存
        )
        
        # 使用正确的方法调用爬虫
        async with AsyncWebCrawler() as crawler:
            result = await crawler.arun(url=main_url, config=crawler_config)
            
            if not result or not result.success or not result.html:
                return None, None, link_counts
                
            print(f"✅ 爬取主页成功: {main_url}")
            
            # 优化：加载历史链接数据并预处理为高效的查找结构
            links_history = load_links_history(source)
            
            # 使用集合存储已知URL，实现O(1)查找
            known_urls_set = set()
            invalid_urls_set = set()
            processed_urls_set = set()
            
            # 预处理链接，只遍历一次历史记录
            for url, data in links_history.items():
                # 将URL规范化存储在集合中
                norm_url = normalize_url(url)
                known_urls_set.add(norm_url)
                
                # 分类存储到不同的集合中实现O(1)查找
                if not data.get("is_valid", False):
                    invalid_urls_set.add(norm_url)
                elif data.get("content_length", 0) > 0:
                    processed_urls_set.add(norm_url)
                    
            print(f"⚠️ 已加载 {len(known_urls_set)} 个已知链接，其中 {len(invalid_urls_set)} 个无效链接")
            print(f"⚠️ 已有 {len(processed_urls_set)} 个链接已经处理过内容")
            
            # 优化：解析HTML并使用字典存储链接，避免重复处理
            soup = BeautifulSoup(result.html, 'html.parser')
            a_tags = soup.find_all('a', href=True)
            
            # 使用字典存储链接，自动去重
            seen_links = {}
            # 优化：使用缓存减少normalize_url重复调用
            normalize_cache = {}
            
            # 第一次遍历：构建规范化链接映射，实现更高效的URL处理
            for a in a_tags:
                href = a['href']
                # 规范化链接，确保完整URL
                if href.startswith('/'):
                    # 相对URL，添加域名
                    parsed_url = urlparse(main_url)
                    base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
                    url = urljoin(base_url, href)
                elif not href.startswith(('http://', 'https://')):
                    # 其他相对URL形式
                    url = urljoin(main_url, href)
                else:
                    url = href
                
                # 使用缓存避免重复规范化
                if url in normalize_cache:
                    norm_url = normalize_cache[url]
                else:
                    norm_url = normalize_url(url)
                    normalize_cache[url] = norm_url
                    
                if not norm_url:
                    continue
                    
                # 使用规范化URL作为键，避免重复
                if norm_url not in seen_links:
                    seen_links[norm_url] = {
                        'url': url,  # 保留原始URL用于显示
                        'a_tag': a,
                        'norm_url': norm_url  # 存储规范化URL用于比较
                    }
            
            print(f"🔍 找到 {len(seen_links)} 个原始链接")
            
            # 需要AI验证的链接
            links_to_validate = []
            
            # 用于储存链接和相关信息
            links_info = []
            
            # 统计重复链接数量
            duplicate_count = 0
            
            # 批量处理链接以提高效率
            
            # 1. 预先提取所有需要处理的链接
            links_to_process = []
            for norm_url, link_data in seen_links.items():
                # 快速检查是否已知链接(使用O(1)复杂度的集合操作)
                if norm_url in known_urls_set:
                    duplicate_count += 1
                    if norm_url in invalid_urls_set:
                        link_counts["skipped_invalid"] += 1
                    elif norm_url in processed_urls_set:
                        link_counts["skipped_processed"] += 1
                    else:
                        link_counts["skipped_existing"] += 1
                    continue
                
                # 只处理新链接
                links_to_process.append((norm_url, link_data))
            
            # 2. 批量处理新链接
            if links_to_process:
                print(f"🆕 发现 {len(links_to_process)} 个新链接，开始处理...")
                
                # 冷启动模式与普通模式的处理逻辑分离，避免每次链接都判断模式
                if USE_COLD_START:
                    # 冷启动模式：批量处理所有新链接
                    batch_updates = {}
                    current_time = datetime.now().isoformat()
                    
                    for norm_url, link_data in links_to_process:
                        url = link_data['url']
                        a_tag = link_data['a_tag']
                        link_text = a_tag.get_text(strip=True) if a_tag else ""
                        
                        link_counts["found_links"] += 1
                        link_counts["new_links"] += 1
                        
                        # 将链接及信息直接添加到结果列表
                        links_info.append({
                            "url": url,
                            "a_tag": a_tag,
                            "title": link_text or url,
                            "is_valid": True,  # 冷启动模式下，默认认为有效
                            "ai_score": 0,
                            "ai_reason": "冷启动模式：跳过验证"
                        })
                        
                        # 准备批量更新数据
                        batch_updates[url] = {
                            "url": url,
                            "title": link_text or url,
                            "first_seen": current_time,
                            "last_updated": current_time,
                            "is_valid": True,
                            "quality_score": 0,
                            "content_length": 0,
                            "error_message": "",
                            "content_fingerprint": "",
                            "ai_score": 0,
                            "ai_reason": "冷启动模式：跳过验证",
                            "crawl_count": 0,
                            "link_type": ""
                        }
                    
                    # 批量更新历史记录
                    if batch_updates:
                        links_history.update(batch_updates)
                        save_links_history(source, links_history)
                        print(f"冷启动模式：已成功将 {len(batch_updates)} 个链接添加到历史记录。")
                    
                    return None, None, link_counts
                else:
                    # 普通模式：批量验证过滤后的链接
                    # 1. 先用基本规则预筛选
                    for norm_url, link_data in links_to_process:
                        url = link_data['url']
                        a_tag = link_data['a_tag']
                        
                        # 计入找到的链接总数
                        link_counts["found_links"] += 1
                        link_counts["new_links"] += 1
                        
                        # 使用基础规则进行初筛
                        is_valid_by_basic = is_valid_news_link(url, main_url, a_tag)
                        
                        # 无论是否通过基础规则，都记录所有新链接到历史记录中
                        current_time = datetime.now().isoformat()
                        links_history[url] = {
                            "url": url,
                            "title": a_tag.get_text(strip=True) if a_tag else url,
                            "first_seen": current_time,
                            "last_updated": current_time,
                            "is_valid": is_valid_by_basic,  # 初始用基本规则判断
                            "quality_score": 0,
                            "content_length": 0,
                            "error_message": "" if is_valid_by_basic else "未通过基本规则检查",
                            "content_fingerprint": "",
                            "ai_score": 0,
                            "ai_reason": "",
                            "crawl_count": 0,
                            "link_type": ""
                        }
                        
                        # 只将通过基础规则的链接添加到待验证列表
                        if is_valid_by_basic:
                            links_to_validate.append({
                                "url": url,
                                "a_tag": a_tag,
                                "title": a_tag.get_text(strip=True) if a_tag else url
                            })
                    
                    # 2. 批量AI验证
                    print(f"🔄 处理后得到 {link_counts['found_links']} 个链接，其中 {len(links_to_validate)} 个需要验证")
                    print(f"⏩ 跳过了 {duplicate_count} 个重复链接")
                    print(f"⏩ 跳过了 {link_counts['skipped_existing']} 个已知有效链接")
                    print(f"⏩ 跳过了 {link_counts['skipped_invalid']} 个已知无效链接")
                    print(f"⏩ 跳过了 {link_counts['skipped_processed']} 个已处理过内容的链接")
                    
                    valid_links = []
                    if links_to_validate:
                        print(f"🧠 使用AI验证 {len(links_to_validate)} 个链接...")
                        
                        # 更新计数
                        link_counts["total_ai_validated"] = len(links_to_validate)
                        
                        if USE_AI_LINK_VALIDATION and AI_LINK_VALIDATOR_AVAILABLE:
                            try:
                                # 导入batch_link_validation函数
                                from ai_link_validator import batch_link_validation
                                
                                # 批量验证链接
                                valid_links = batch_link_validation(links_to_validate, main_url, result.html)
                                
                                # 调试打印验证结果
                                print(f"✅ AI验证完成，得到 {len(valid_links)} 个有效链接")
                                
                                # 更新结果统计
                                link_counts["valid_links"] = len(valid_links)
                                link_counts["invalid_links"] = len(links_to_validate) - len(valid_links)
                                
                                # 将验证结果添加到链接信息列表
                                for link_info in valid_links:
                                    link_info["is_valid"] = True
                                    links_info.append(link_info)
                                    
                                    # 立即将有效链接写入Excel
                                    write_valid_link_to_excel(
                                        url=link_info.get("url"),
                                        title=link_info.get("title", ""),
                                        source_name=source_name,
                                        ai_score=link_info.get("ai_score", 0),
                                        ai_reason=link_info.get("ai_reason", ""),
                                        link_type=link_info.get("link_type", "")
                                    )
                                
                                # 优化：批量更新历史记录 - 更高效的实现
                                batch_updates = {}
                                current_time = datetime.now().isoformat()
                                validated_urls = {normalize_url(info.get("url")): info for info in valid_links}
                                
                                # 一次性准备所有更新数据 - 确保所有链接都被更新到历史记录中
                                for link in links_to_validate:
                                    url = link.get("url")
                                    norm_url = normalize_url(url)
                                    
                                    # 检查是否通过验证
                                    is_valid = norm_url in validated_urls
                                    link_info = validated_urls.get(norm_url, {})
                                    
                                    ai_score = link_info.get("ai_score", 0) if is_valid else 0
                                    ai_reason = link_info.get("ai_reason", "") if is_valid else "AI验证未通过"
                                    
                                    # 准备更新数据 - 无论是否有效都会更新
                                    batch_updates[url] = {
                                        "url": url,
                                        "title": link.get("title", ""),
                                        "first_seen": current_time,
                                        "last_updated": current_time,
                                        "is_valid": is_valid,
                                        "quality_score": 0,
                                        "content_length": 0,
                                        "error_message": "" if is_valid else "AI验证未通过",
                                        "content_fingerprint": "",
                                        "ai_score": ai_score,
                                        "ai_reason": ai_reason,
                                        "crawl_count": 0,
                                        "link_type": link_info.get("link_type", "")
                                    }
                                
                                # 批量更新历史记录
                                links_history.update(batch_updates)
                                save_links_history(source, links_history)
                                print(f"✅ 已更新 {len(batch_updates)} 个链接到历史记录")
                                
                            except Exception as e:
                                print(f"❌ AI批量验证失败: {str(e)}")
                                # 对于异常情况，仍然确保链接被添加到历史记录
                                batch_updates = {}
                                current_time = datetime.now().isoformat()
                                
                                for link in links_to_validate:
                                    url = link.get("url")
                                    batch_updates[url] = {
                                        "url": url,
                                        "title": link.get("title", ""),
                                        "first_seen": current_time,
                                        "last_updated": current_time,
                                        "is_valid": True,  # 出错时默认为有效
                                        "quality_score": 0,
                                        "content_length": 0,
                                        "error_message": f"AI验证失败：{str(e)}",
                                        "content_fingerprint": "",
                                        "ai_score": 0,
                                        "ai_reason": f"AI验证失败，默认视为有效: {str(e)}",
                                        "crawl_count": 0,
                                        "link_type": ""
                                    }
                                
                                links_history.update(batch_updates)
                                save_links_history(source, links_history)
                                print(f"✅ 已更新 {len(batch_updates)} 个链接到历史记录（AI验证失败）")
                        else:
                            # 不使用AI验证，所有通过基础规则的链接都视为有效
                            for link in links_to_validate:
                                link["is_valid"] = True
                                link["ai_score"] = 0
                                link["ai_reason"] = "跳过AI验证"
                                links_info.append(link)
                                valid_links.append(link)
                            
                            link_counts["valid_links"] = len(valid_links)
                    else:
                        print("⚠️ 没有需要验证的新链接")
                    
                    # 在处理结束后再次保存历史记录（确保所有链接都被记录）
                    save_links_history(source, links_history)
                    
                    return valid_links, result.html, link_counts
            else:
                print("⚠️ 没有发现新链接")
                return [], result.html, link_counts
            
    except Exception as e:
        print(f"❌ 获取链接失败: {str(e)}")
        traceback.print_exc()
        return None, None, link_counts

# 导出历史链接库到Excel（新增函数）
def export_links_history_to_excel():
    """将所有历史链接数据导出到Excel文件"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 确保文件保存在当前目录下
        current_dir = os.path.dirname(os.path.abspath(__file__))
        filename = os.path.join(current_dir, f"links_history_{timestamp}.xlsx")
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "历史链接"
        
        # 设置表头
        headers = [
            '序号', '链接ID', '来源', 'URL', '标题', '有效性', '首次发现时间', 
            '最后更新时间', '内容长度', '质量评分', '爬取次数', 'AI验证分数', 'AI验证理由', '链接类型'
        ]
        
        # 设置列宽
        column_widths = {
            'A': 8,   # 序号
            'B': 35,  # 链接ID
            'C': 15,  # 来源
            'D': 60,  # URL
            'E': 50,  # 标题
            'F': 10,  # 有效性
            'G': 20,  # 首次发现时间
            'H': 20,  # 最后更新时间
            'I': 12,  # 内容长度
            'J': 12,  # 质量评分
            'K': 10,  # 爬取次数
            'L': 12,  # AI验证分数
            'M': 40,  # AI验证理由
            'N': 15,  # 链接类型
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # 写入表头
        for idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=idx).value = header
        
        # 遍历所有历史记录文件
        row_idx = 2
        total_records = 0
        
        # 确保目录存在
        if not os.path.exists(LINKS_HISTORY_DIR):
            print(f"历史链接目录不存在: {LINKS_HISTORY_DIR}")
            return None
            
        # 获取所有历史记录文件
        history_files = [f for f in os.listdir(LINKS_HISTORY_DIR) if f.endswith('_links.json')]
        
        for history_file in history_files:
            # 提取来源ID
            source = history_file.replace('_links.json', '')
            
            # 加载数据
            file_path = os.path.join(LINKS_HISTORY_DIR, history_file)
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    links_data = json.load(f)
                    
                # 写入数据
                for link_id, info in links_data.items():
                    # 构建数据行
                    data = [
                        row_idx - 1,  # 序号
                        link_id,      # 链接ID
                        source,       # 来源
                        info.get('url', ''),            # URL
                        info.get('title', ''),          # 标题
                        '有效' if info.get('is_valid', False) else '无效',  # 有效性
                        info.get('first_seen', ''),     # 首次发现时间
                        info.get('last_updated', ''),   # 最后更新时间
                        info.get('content_length', 0),  # 内容长度
                        info.get('quality_score', 0),   # 质量评分
                        info.get('crawl_count', 0),     # 爬取次数
                        info.get('ai_score', 0),        # AI验证分数
                        info.get('ai_reason', ''),      # AI验证理由
                        info.get('link_type', '')       # 链接类型
                    ]
                    
                    # 写入一行数据
                    for col, value in enumerate(data, 1):
                        sheet.cell(row=row_idx, column=col).value = value
                        
                    row_idx += 1
                    total_records += 1
                    
            except Exception as e:
                print(f"处理历史文件时出错 {history_file}: {e}")
                continue
        
        # 保存文件
        wb.save(filename)
        print(f"成功导出历史链接数据到 {filename}，共 {total_records} 条记录")
        return filename
        
    except Exception as e:
        print(f"导出历史链接数据时出错: {e}")
        return None

# 将抓取结果写入Excel文件
def write_result_to_excel(url, title, source_name, publish_date, content, link_duration, 
                          is_valid_content, retry_count, processed_count, ai_score=0, 
                          ai_reason="", link_type=""):
    """将抓取结果写入Excel文件"""
    if not SAVE_TO_EXCEL or not result_excel_file or not os.path.exists(result_excel_file):
        return
    
    try:
        # 加载Excel文件
        wb = openpyxl.load_workbook(result_excel_file)
        sheet = wb.active
        next_row = sheet.max_row + 1
        
        # 构建数据
        data = [
            processed_count,  # 索引
            source_name,      # 来源
            url,              # 链接
            title,            # 标题
            publish_date,     # 发布日期
            "新闻" if is_valid_content else "非新闻",  # 分类
            len(content),     # 正文字数
            f"{link_duration:.2f}",    # 爬取时间
            "成功" if is_valid_content else "失败",  # 状态
            retry_count,      # 尝试次数
            ai_score,         # AI验证分数
            ai_reason,        # AI验证理由
            link_type         # 链接类型
        ]
        
        # 写入数据
        for col, value in enumerate(data, 1):
            sheet.cell(row=next_row, column=col).value = value
            
        # 修改：立即保存Excel文件
        wb.save(result_excel_file)
        print(f"✅ 已将结果写入Excel: 行 {next_row} 并保存")
    except Exception as e:
        print(f"❌ 写入Excel时出错: {e}")

# 修改写入Excel函数，确保实时保存，并在每个有效链接被验证后立即写入Excel
def write_valid_link_to_excel(url, title, source_name, ai_score=0, ai_reason="", link_type=""):
    """将有效链接信息立即写入Excel文件"""
    if not SAVE_TO_EXCEL:
        return
    
    # 确保result_excel_file是全局变量
    global result_excel_file
    
    # 如果结果文件不存在，先创建
    if not result_excel_file or not os.path.exists(result_excel_file):
        result_excel_file = create_result_excel()
        if not result_excel_file:
            print(f"❌ 无法创建Excel结果文件")
            return
    
    try:
        # 加载Excel文件
        wb = openpyxl.load_workbook(result_excel_file)
        sheet = wb.active
        next_row = sheet.max_row + 1
        
        # 构建基本数据
        data = [
            next_row-1,     # 索引
            source_name,    # 来源
            url,            # 链接
            title,          # 标题
            "待处理",       # 发布日期（待爬取内容后更新）
            "新闻",         # 分类
            0,              # 正文字数（待爬取内容后更新）
            0,              # 爬取时间（待爬取内容后更新）
            "待处理",       # 状态
            0,              # 尝试次数
            ai_score,       # AI验证分数
            ai_reason,      # AI验证理由
            link_type       # 链接类型
        ]
        
        # 写入数据
        for col, value in enumerate(data, 1):
            sheet.cell(row=next_row, column=col).value = value
        
        # 立即保存Excel文件
        wb.save(result_excel_file)
        print(f"✅ 已将有效链接写入Excel: 行 {next_row} 并保存")
    except Exception as e:
        print(f"❌ 写入Excel时出错: {e}")
        traceback.print_exc()

# 主程序
async def main():
    """主函数入口"""
    # 创建结果Excel文件
    if SAVE_TO_EXCEL:
        result_file = create_result_excel()
        print(f"📊 Excel结果将保存到: {result_file}")
    
    # 设置批次ID（用于文件名和日志）
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(f"🚀 开始新闻爬取，批次ID: {batch_id}")
    
    # 读取excel文件中的主页链接 - 使用当前脚本目录下的文件
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_source_file = os.path.join(current_dir, "testhomepage.xlsx")
    
    # 检查文件是否存在
    if not os.path.exists(excel_source_file):
        print(f"⚠️ 源文件不存在: {excel_source_file}")
        excel_source_file = "../testhomepage.xlsx"
        if not os.path.exists(excel_source_file):
            print(f"⚠️ 备用源文件也不存在: {excel_source_file}")
            # 尝试列出当前目录和上级目录中的Excel文件
            print("当前目录下的Excel文件:")
            for file in os.listdir('.'):
                if file.endswith('.xlsx'):
                    print(f" - {file}")
            
            print("上级目录下的Excel文件:")
            for file in os.listdir('..'):
                if file.endswith('.xlsx'):
                    print(f" - {file}")
            return
    
    print(f"📄 找到源文件: {excel_source_file}")
    wb_source = openpyxl.load_workbook(excel_source_file)
    sheet_source = wb_source.active
    
    # 创建历史链接目录
    os.makedirs(LINKS_HISTORY_DIR, exist_ok=True)
    
    # 记录统计信息
    source_count = 0         # 处理的来源数量
    processed_count = 0      # 处理的链接数量
    success_count = 0        # 成功处理的链接数量
    error_count = 0          # 处理失败的链接数量
    skipped_count = 0        # 跳过的链接数量
    new_link_count = 0       # 新链接数量
    invalid_link_count = 0   # 无效链接数量
    
    # 遍历excel文件中的每一行
    for row in range(2, sheet_source.max_row + 1):  # 从第2行开始，跳过标题行
        try:
            remark = sheet_source.cell(row=row, column=1).value
            main_url = sheet_source.cell(row=row, column=2).value
            source_name = sheet_source.cell(row=row, column=3).value
            
            if not main_url or not source_name:
                continue
                
            source_count += 1
            source_success_count = 0  # 初始化每个来源的成功计数器
            
            # 每个源的统计信息
            print(f"\n🌐 处理来源 {source_count}/{source_count}: {source_name} - {main_url} (Source ID: {source_name})")
            
            # 加载此源的历史链接
            links_history_file = os.path.join(LINKS_HISTORY_DIR, f"{source_name}_links.json")
            print(f"📋 历史链接文件: {links_history_file}")
            
            # 爬取主页中的新闻链接
            start_time = time.time()
            links, html_content, stats = await fetch_news_links(main_url, source_name)
            end_time = time.time()
            
            print(f"链接爬取时间: {end_time - start_time:.2f} 秒")
            print(f"找到 {len(links) if links else 0} 个链接需要验证")
            
            # 打印链接过滤详情
            print("链接过滤详情:")
            print(f"- 原始链接总数: {stats['found_links']}")
            print(f"- 跳过重复链接: {stats['skipped_existing']}")
            print(f"- 跳过已知无效链接: {stats['skipped_invalid']}")
            print(f"- 跳过了 {stats['skipped_processed']} 个已处理过内容的链接")
            
            # 如果是冷启动模式，直接返回结果，不进行实际内容爬取
            if USE_COLD_START:
                print("冷启动模式：跳过内容爬取步骤")
                # 在循环中不要直接返回，而是记录结果并继续处理下一个来源
                # return {
                #     'source': source_name,
                #     'url': main_url,
                #     'valid_links': len(all_processed_links),
                #     'message': f"冷启动模式：已将 {len(all_processed_links)} 个链接添加到历史记录"
                # }
                # 直接继续处理下一个来源
                continue
            
            # 进行AI验证
            valid_links = []
            
            if USE_AI_LINK_VALIDATION and AI_LINK_VALIDATOR_AVAILABLE and links:
                # 确保链接对象结构一致
                links_for_validation = [
                    {'url': item.get('url', item.get('link')), 'title': item.get('title', '')}
                    for item in links
                ]
                
                # 检查是否有有效的链接需要验证
                if links_for_validation:
                    # 使用AI验证链接，这里的html内容可以为空
                    crawler_html = ""  # 我们不再依赖result变量
                    try:
                        from ai_link_validator import batch_link_validation
                        validated_links = batch_link_validation(
                            links_for_validation, 
                            main_url, 
                            crawler_html
                        )
                        
                        # 将验证后的链接与原始链接合并，保持原始的link_obj结构
                        # 创建一个URL到验证结果的映射
                        validation_results = {}
                        for validated_link in validated_links:
                            validation_results[validated_link.get('url')] = validated_link
                        
                        # 更新原始链接的验证信息
                        for i, original_link in enumerate(links):
                            # 确定URL键
                            url_key = 'url' if 'url' in original_link else 'link'
                            link_url = original_link.get(url_key)
                            
                            if link_url in validation_results:
                                validated_info = validation_results[link_url]
                                # 更新验证信息
                                original_link['is_valid'] = validated_info.get('is_valid', False)
                                original_link['ai_score'] = validated_info.get('ai_score', 0)
                                original_link['ai_reason'] = validated_info.get('ai_reason', '')
                                original_link['link_type'] = validated_info.get('link_type', '')
                                
                                # 如果验证结果是有效的，添加到valid_links
                                if validated_info.get('is_valid', False):
                                    # 将有效链接添加到valid_links
                                    valid_links.append(original_link)
                                    
                                    # 立即将有效链接写入Excel
                                    write_valid_link_to_excel(
                                        url=link_url,
                                        title=original_link.get('title', ''),
                                        source_name=source_name,
                                        ai_score=validated_info.get('ai_score', 0),
                                        ai_reason=validated_info.get('ai_reason', ''),
                                        link_type=validated_info.get('link_type', '')
                                    )
                    except Exception as e:
                        print(f"❌ AI验证过程出错: {str(e)}")
                        print("⚠️ 跳过AI验证，所有链接将被视为有效")
                        # 默认所有链接有效
                        valid_links = links
                        
                        # 确保即使出错也将链接写入Excel
                        for link in links:
                            url_key = 'url' if 'url' in link else 'link'
                            url = link.get(url_key)
                            title = link.get('title', '')
                            
                            # 立即将链接写入Excel
                            write_valid_link_to_excel(
                                url=url,
                                title=title,
                                source_name=source_name,
                                ai_score=0,
                                ai_reason="AI验证失败，默认视为有效",
                                link_type=""
                            )
                else:
                    print("⚠️ 没有有效链接需要验证")
            else:
                # 没有使用AI验证时，所有链接都视为有效
                valid_links = links
                
                # 确保所有链接都被写入Excel
                for link in links:
                    url_key = 'url' if 'url' in link else 'link'
                    url = link.get(url_key)
                    title = link.get('title', '')
                    
                    # 立即将链接写入Excel
                    write_valid_link_to_excel(
                        url=url,
                        title=title,
                        source_name=source_name,
                        ai_score=0,
                        ai_reason="跳过AI验证",
                        link_type=""
                    )
            
            # 更新链接列表为验证后的有效链接
            links = valid_links
            
            # 记录开始时间
            start_content_time = time.time()
            
            # 处理每个链接的内容
            for link_obj in links:
                # 健壮性处理：确保link_obj具有url
                if isinstance(link_obj, dict):
                    # 有些link_obj可能使用'url'键，有些可能使用'link'键
                    link = link_obj.get('url')
                    if link is None:
                        link = link_obj.get('link')
                    
                    title = link_obj.get('title', link if link else '')
                else:
                    # 如果link_obj不是字典类型
                    print(f"⚠️ 未知的链接对象类型: {type(link_obj)}，跳过处理")
                    continue
                
                if not link:
                    print("⚠️ 链接为空，跳过处理")
                    continue
                
                processed_count += 1
                print(f"\n处理链接 {processed_count}/{len(links)}: {link}")
                
                # 检查链接是否已经被处理过
                is_new, is_invalid, is_processed = is_new_link(link, source_name)
                
                if not is_new and is_processed:
                    print(f"跳过已处理过的链接: {link}")
                    continue
                
                # 记录是否为新链接
                if is_new:
                    new_link_count += 1
                
                retry_count = 0
                
                # 只有判定为有效链接才进行内容爬取
                if link_obj.get('is_valid', True):
                    while retry_count <= MAX_RETRY_COUNT:
                        try:
                            is_new, _, is_processed = is_new_link(link, source_name)
                            
                            if not is_new and is_processed:
                                print(f"跳过已处理过的链接: {link}")
                                break
                                
                            crawl_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            
                            print(f"\n🔗 正在处理链接: {link}")
                            print(f"📌 链接标题: {title}")
                            print(f"🆕 是否新链接: {'是' if is_new else '否'}")
                            
                            # 设置浏览器配置
                            browser_config = BrowserConfig(
                                browser_type="chromium",
                                headless=True, 
                                viewport_width=1366,
                                viewport_height=768,
                                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36",
                                java_script_enabled=True,
                                ignore_https_errors=True
                            )
                            
                            # 设置爬虫配置
                            prune_filter = PruningContentFilter(
                                threshold=0.45,
                                threshold_type="dynamic",
                                min_word_threshold=5
                            )
                            md_generator = DefaultMarkdownGenerator(content_filter=prune_filter)
                            
                            crawler_config = CrawlerRunConfig(
                                markdown_generator=md_generator,
                                page_timeout=REQUEST_TIMEOUT * 1000,
                                cache_mode=CacheMode.BYPASS
                            )
                            
                            # 爬取内容
                            link_start_time = time.time()
                            async with AsyncWebCrawler(config=browser_config) as crawler:
                                result = await crawler.arun(url=link, config=crawler_config)
                            link_end_time = time.time()
                            
                            # 计算爬取耗时
                            link_duration = link_end_time - link_start_time
                            
                            if result.success:
                                print(f"✅ 成功获取内容, 耗时: {link_duration:.2f}秒")
                                
                                # 使用BeautifulSoup处理HTML
                                soup = BeautifulSoup(result.html, 'html.parser')
                                title = extract_title(soup, link) or title
                                publish_date = extract_publish_date(soup) or "未找到日期"
                                content = result.markdown.fit_markdown
                                content_length = len(content)
                                
                                print(f"📝 标题: {title}")
                                print(f"📅 发布日期: {publish_date}")
                                print(f"📊 内容长度: {content_length} 字符")
                                
                                # 使用evaluate_content_quality获取摘要和指纹（不再评估内容质量）
                                _, content_summary, content_fingerprint = evaluate_content_quality(result.html, title, link)
                                
                                # AI验证通过的链接都视为有效
                                is_valid_content = True
                                
                                # 输出结果
                                print(f"✅ 成功获取内容, 耗时: {link_duration:.2f}秒")
                                print(f"💡 内容摘要:\n{content_summary}")
                                
                                # 爬取内容后，更新链接历史记录，设置内容长度以标记为已处理
                                link_obj['is_processed'] = True
                                
                                # 获取AI验证信息
                                ai_score = link_obj.get('ai_score', 0)
                                ai_reason = link_obj.get('ai_reason', "")
                                link_type = link_obj.get('link_type', "")
                                
                                # 更新链接历史
                                update_link_history(
                                    url=link,
                                    title=title,
                                    source=source_name,
                                    content_summary=content_summary,
                                    is_valid=is_valid_content,
                                    quality_score=ai_score,
                                    content_length=content_length,
                                    content_fingerprint=content_fingerprint,
                                    ai_score=ai_score,
                                    ai_reason=ai_reason,
                                    link_type=link_type
                                )
                                
                                # 写入Excel
                                if SAVE_TO_EXCEL:
                                    write_result_to_excel(
                                        url=link,
                                        title=title,
                                        source_name=source_name,
                                        publish_date=publish_date,
                                        content=content,
                                        link_duration=link_duration,
                                        is_valid_content=is_valid_content,
                                        retry_count=retry_count,
                                        processed_count=processed_count,
                                        ai_score=ai_score,
                                        ai_reason=ai_reason,
                                        link_type=link_type
                                    )
                                
                                source_success_count += 1
                                success_count += 1
                                break  # 成功，跳出重试循环
                                
                            else:
                                print(f"❌ 获取内容失败: {result.error_message}")
                                error_message = result.error_message
                                
                                # 更新链接历史记录 - 标记为获取失败
                                update_link_history(
                                    url=link,
                                    title=title,
                                    source=source_name,
                                    is_valid=False,
                                    error_message=f"爬取失败: {error_message}"
                                )
                                
                                retry_count += 1
                                if retry_count <= MAX_RETRY_COUNT:
                                    print(f"⏳ 重试 ({retry_count}/{MAX_RETRY_COUNT})...")
                                else:
                                    print(f"❌ 达到最大重试次数，放弃此链接")
                                    error_count += 1
                                
                        except Exception as e:
                            print(f"❌ 处理链接时出错: {e}")
                            traceback.print_exc()
                            
                            # 更新链接历史记录 - 标记为处理异常
                            update_link_history(
                                url=link,
                                title=title,
                                source=source_name,
                                is_valid=False,
                                error_message=f"处理异常: {str(e)}"
                            )
                            
                            retry_count += 1
                            if retry_count <= MAX_RETRY_COUNT:
                                print(f"⏳ 重试 ({retry_count}/{MAX_RETRY_COUNT})...")
                            else:
                                print(f"❌ 达到最大重试次数，放弃此链接")
                                error_count += 1
                
                # 完成当前来源的处理
                print(f"\n✅ 来源 {source_name} 处理完成，成功获取 {source_success_count} 个有效链接")
                
        except Exception as e:
            print(f"❌ 处理源 {source_name} 时出错: {str(e)}")
            traceback.print_exc()
            continue
            
    # 详细的统计信息打印
    print(f"\n📊 爬取统计信息:")
    print(f"总共处理了 {source_count} 个来源")
    print(f"总共发现了 {processed_count + skipped_count} 个链接")
    print(f" - 其中 {new_link_count} 个是新链接")
    print(f" - 其中 {skipped_count} 个链接被跳过（已处理过或已知无效）")
    print(f"实际处理了 {processed_count} 个链接")
    print(f" - 其中 {success_count} 个有效并成功爬取")
    print(f" - 其中 {error_count} 个处理失败")
    print(f" - 其中 {invalid_link_count} 个被AI判断为无效")
    
    # 计算有效率
    if processed_count > 0:
        success_rate = (success_count / processed_count) * 100
        print(f"有效率: {success_rate:.2f}%")
    
    # 导出历史链接到Excel
    history_file = export_links_history_to_excel()
    if history_file:
        print(f"📊 历史链接已导出到: {history_file}")
    
    print(f"🏁 程序结束，批次ID: {batch_id}")

# 当直接运行此脚本时执行主函数
if __name__ == "__main__":
    import sys
    
    # 检查命令行参数
    if len(sys.argv) > 1 and sys.argv[1] == "--debug":
        # 调试模式：爬取单个链接
        if len(sys.argv) >= 4:
            # 格式: python test_request3.py --debug <url> <source_name>
            debug_url = sys.argv[2]
            debug_source = sys.argv[3]
            
            print(f"🔍 调试模式: 爬取单个链接")
            print(f"🔗 URL: {debug_url}")
            print(f"📌 来源: {debug_source}")
            
            # 创建结果Excel文件
            if SAVE_TO_EXCEL:
                result_file = create_result_excel()
                print(f"📊 Excel结果将保存到: {result_file}")
            
            # 确保历史链接目录存在
            os.makedirs(LINKS_HISTORY_DIR, exist_ok=True)
            
            # 定义单个链接调试函数
            async def debug_single_link():
                try:
                    # 创建爬虫配置
                    browser_config = BrowserConfig(
                        browser_type="chromium",
                        viewport_width=1280, 
                        viewport_height=800,
                        java_script_enabled=True,
                        ignore_https_errors=True
                    )
                    
                    # 设置过滤器
                    prune_filter = PruningContentFilter(
                        threshold=0.45,
                        threshold_type="dynamic",
                        min_word_threshold=5
                    )
                    
                    # 设置Markdown生成器
                    md_generator = DefaultMarkdownGenerator(content_filter=prune_filter)
                    
                    # 爬虫配置
                    crawler_config = CrawlerRunConfig(
                        markdown_generator=md_generator,
                        page_timeout=REQUEST_TIMEOUT * 1000,
                        cache_mode=CacheMode.BYPASS
                    )
                    
                    # 先验证链接是否为有效新闻链接
                    print(f"🧠 使用AI验证链接有效性...")
                    parsed = urlparse(debug_url)
                    base_url = f"{parsed.scheme}://{parsed.netloc}"
                    
                    if AI_LINK_VALIDATOR_AVAILABLE:
                        # 使用AI验证
                        is_valid = is_valid_news_link_with_ai(debug_url, base_url)
                        if not is_valid:
                            print(f"❌ AI判断该链接不是有效的新闻链接，但仍将尝试爬取内容")
                    else:
                        print(f"⚠️ AI验证模块不可用，将直接爬取内容")
                    
                    print(f"🔄 开始爬取链接: {debug_url}")
                    start_time = time.time()
                    
                    # 创建爬虫实例并爬取内容
                    async with AsyncWebCrawler() as crawler:
                        result = await crawler.arun(url=debug_url, config=crawler_config)
                    
                    end_time = time.time()
                    duration = end_time - start_time
                    
                    if result.success:
                        # 解析内容
                        soup = BeautifulSoup(result.html, 'html.parser')
                        title = extract_title(soup, debug_url)
                        publish_date = extract_publish_date(soup) or "未找到日期"
                        content = result.markdown.fit_markdown if hasattr(result.markdown, 'fit_markdown') else ""
                        content_length = len(content)
                        
                        # 使用evaluate_content_quality获取摘要和指纹（不再评估内容质量）
                        _, content_summary, content_fingerprint = evaluate_content_quality(result.html, title, debug_url)
                        
                        # 获取AI验证信息
                        ai_score = 0
                        ai_reason = ""
                        link_type = ""
                        
                        # 从AI验证日志中查找此链接的记录
                        if ENABLE_LOGGING and os.path.exists(LOG_FILE):
                            try:
                                with open(LOG_FILE, "r", encoding="utf-8") as f:
                                    for line in f:
                                        try:
                                            record = json.loads(line)
                                            if record.get('url') == debug_url:
                                                ai_score = record.get('score', 0)
                                                ai_reason = record.get('reason', "")
                                                link_type = "文章" if record.get('is_valid', False) else "非文章"
                                                break
                                        except json.JSONDecodeError:
                                            continue
                            except Exception as e:
                                print(f"读取AI验证日志时出错: {e}")
                        
                        # 输出结果
                        print(f"\n✅ 爬取成功! 耗时: {duration:.2f}秒")
                        print(f"📝 标题: {title}")
                        print(f"📅 发布日期: {publish_date}")
                        print(f"📊 内容长度: {content_length}字符")
                        if ai_score > 0:
                            print(f"🧠 AI验证分数: {ai_score}/100")
                            print(f"🧠 AI验证理由: {ai_reason}")
                        print(f"💡 内容摘要:\n{content_summary}")
                        
                        # 更新历史链接
                        update_link_history(
                            url=debug_url,
                            title=title,
                            source=debug_source,
                            content_summary=content_summary,
                            is_valid=True,  # 调试模式下始终视为有效
                            quality_score=ai_score,
                            content_length=content_length,
                            content_fingerprint=content_fingerprint,
                            ai_score=ai_score,
                            ai_reason=ai_reason,
                            link_type=link_type
                        )
                        
                        # 写入Excel结果
                        if SAVE_TO_EXCEL:
                            write_result_to_excel(
                                url=debug_url,
                                title=title,
                                source_name=debug_source,
                                publish_date=publish_date,
                                content=content,
                                link_duration=duration,
                                is_valid_content=True,
                                retry_count=0,
                                processed_count=1,
                                ai_score=ai_score,
                                ai_reason=ai_reason,
                                link_type=link_type
                            )
                        
                        # 打印部分内容
                        print("\n📄 内容预览 (前500字符):")
                        print(f"{content[:500]}...")
                        print(f"\n完整内容已保存到Excel文件: {result_excel_file}")
                        
                    else:
                        print(f"\n❌ 爬取失败! 耗时: {duration:.2f}秒")
                        print(f"错误信息: {result.error_message}")
                
                except Exception as e:
                    print(f"\n⚠️ 调试模式出错: {str(e)}")
                    traceback.print_exc()
            
            # 执行单链接调试
            asyncio.run(debug_single_link())
            
        else:
            print("❌ 调试模式用法: python test_request3.py --debug <url> <source_name>")
            print("例如: python test_request3.py --debug https://example.com/news/article123 source1")
    else:
        # 正常模式：运行完整爬取流程
        asyncio.run(main()) 