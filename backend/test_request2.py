import openpyxl
import asyncio
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig, BrowserConfig
from bs4 import BeautifulSoup
import os
import json
import re
from datetime import datetime
from urllib.parse import urlparse, urljoin
from crawl4ai.extraction_strategy import JsonCssExtractionStrategy
from crawl4ai import CacheMode

# 配置选项
SAVE_TO_EXCEL = True  # 设置为False可以禁用Excel输出功能

# 读取主界面链接
source_file = '../testhomepage.xlsx'
wb = openpyxl.load_workbook(source_file)
sheet = wb.active

# 创建结果Excel文件
result_file = 'news_results.xlsx'
if os.path.exists(result_file):
    result_wb = openpyxl.load_workbook(result_file)
    result_sheet = result_wb.active
else:
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    result_sheet.title = 'News Results'
    headers = ['来源', '子链接', '标题', '发布时间']
    for col, header in enumerate(headers, 1):
        result_sheet.cell(row=1, column=col).value = header

# 读取已存在的子链接以避免重复
existing_links = set()
for row in result_sheet.iter_rows(min_row=2, values_only=True):
    existing_links.add(row[1])

# 判断链接是否为有效新闻链接的函数
def is_valid_news_link(link, base_url):
    """
    判断链接是否为有效的新闻链接
    
    过滤条件:
    1. 排除社交媒体链接
    2. 排除分页链接 (通常包含 'page', 'pagination', 'prev', 'next', 等)
    3. 排除搜索、标签、类别等非新闻内容链接
    4. 排除外部链接（非同一域名的链接）
    5. 排除杂项链接（如登录、注册、关于我们等）
    """
    # 如果链接为空或者是锚点，返回False
    if not link or link.startswith('#') or link.startswith('javascript:'):
        return False
    
    # 解析链接和基础URL
    parsed_link = urlparse(link)
    parsed_base = urlparse(base_url)
    
    # 如果是相对链接，转换为完整URL
    if not parsed_link.netloc:
        full_link = urljoin(base_url, link)
        parsed_link = urlparse(full_link)
    
    # 检查是否为同一域名（排除外部链接）
    if parsed_link.netloc != parsed_base.netloc:
        return False
    
    # 排除社交媒体链接
    social_media_patterns = ['facebook', 'twitter', 'instagram', 'linkedin', 'youtube', 
                           'tiktok', 'telegram', 'pinterest', 'reddit', 'tumblr', 'whatsapp']
    
    if any(sm in parsed_link.netloc.lower() for sm in social_media_patterns):
        return False
    
    # 排除常见的非新闻内容路径
    non_news_paths = ['/about', '/contact', '/terms', '/privacy', '/login', '/register', 
                     '/subscribe', '/search', '/feed', '/rss', '/sitemap', '/tag', 
                     '/category', '/author', '/comment', '/user', '/profile', '/newsletter',
                     '/subscribe', '/advertise', '/page']
    
    if any(path in parsed_link.path.lower() for path in non_news_paths):
        return False
    
    # 排除分页链接
    pagination_patterns = ['/page', 'pagination', 'prev', 'next', 'before', 'after']
    if any(pattern in parsed_link.path.lower() for pattern in pagination_patterns):
        return False
    
    # 排除常见的资源文件后缀
    resource_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.pdf', '.zip', '.doc', 
                          '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.css', '.js']
    if any(parsed_link.path.lower().endswith(ext) for ext in resource_extensions):
        return False
    
    # 检查是否可能是新闻文章的链接特征
    # 常见的新闻URL模式包含 /news/, /article/, /story/, 日期(YYYY/MM/DD), 数字ID等
    news_patterns = ['/news/', '/article/', '/story/', '/post/', '/nota/', '/noticias/', 
                    '/agro/', '/economy/', '/economia/', '/daily/', '/content/',
                    'topics', 'video', 'report', 'analysis']
    
    has_date_pattern = bool(re.search(r'/20\d{2}[/-]\d{1,2}[/-]\d{1,2}/', parsed_link.path))
    has_news_id = bool(re.search(r'/\d{4,}/', parsed_link.path))
    
    if (any(pattern in parsed_link.path.lower() for pattern in news_patterns) or 
        has_date_pattern or has_news_id):
        return True
    
    # 如果路径长度适中（可能是文章），且不包含常见的非新闻路径模式
    path_segments = [seg for seg in parsed_link.path.split('/') if seg]
    if 2 <= len(path_segments) <= 5:
        return True
    
    # 默认情况下返回False
    return False

# 爬取主界面
async def fetch_news_links(main_url, source):
    print(f"\n开始爬取来源 {source}: {main_url}")
    
    # 创建优化的爬虫配置
    browser_config = BrowserConfig(
        headless=True,
        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    )
    
    crawler_config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        page_timeout=60000,  # 设置较长的超时时间
        excluded_tags=["nav", "footer", "header", "aside", "form"],
        exclude_external_links=True  # 排除外部链接
    )
    
    try:
        async with AsyncWebCrawler(config=browser_config) as crawler:
            result = await crawler.arun(url=main_url, config=crawler_config)
            
            if result.success:
                print(f"✅ 成功爬取 {main_url}")
                
                links_found = False
                try:
                    # 使用BeautifulSoup处理HTML
                    soup = BeautifulSoup(result.html, 'html.parser')
                    all_links = soup.find_all('a', href=True)
                    print(f"找到 {len(all_links)} 个链接，开始筛选有效新闻链接...")
                    
                    valid_links_count = 0
                    
                    # 处理所有链接
                    for a_tag in all_links:
                        link = a_tag['href']
                        title = a_tag.get_text(strip=True)
                        
                        # 跳过空标题或者过短的标题
                        if not title or len(title) < 5:
                            continue
                        
                        # 规范化URL
                        if link.startswith('/'):
                            # 相对URL，添加域名
                            parsed_url = urlparse(main_url)
                            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
                            link = urljoin(base_url, link)
                        
                        # 检查是否为有效新闻链接
                        if is_valid_news_link(link, main_url):
                            valid_links_count += 1
                            print(f"  有效新闻链接: {link}")
                            print(f"  标题: {title}")
                            
                            # 检查链接是否已存在
                            if link not in existing_links:
                                # 准备爬取详情页以获取发布时间
                                try:
                                    publish_time = "未知"
                                    
                                    # 爬取新闻详情页
                                    detail_result = await crawler.arun(
                                        url=link, 
                                        config=CrawlerRunConfig(
                                            cache_mode=CacheMode.BYPASS
                                        )
                                    )
                                    
                                    if detail_result.success:
                                        # 尝试从HTML中提取发布时间
                                        detail_soup = BeautifulSoup(detail_result.html, 'html.parser')
                                        
                                        # 查找常见的元数据标签
                                        meta_tags = detail_soup.find_all('meta')
                                        
                                        # 检查各种可能的日期元标签
                                        date_properties = [
                                            'article:published_time', 'datePublished', 
                                            'publishedDate', 'pubdate', 'date', 'DC.date',
                                            'article:modified_time', 'lastModified'
                                        ]
                                        
                                        for meta in meta_tags:
                                            if meta.get('property') in date_properties or meta.get('name') in date_properties:
                                                if meta.get('content'):
                                                    publish_time = meta.get('content')
                                                    break
                                        
                                        # 如果元标签中没有日期，查找常见的日期容器
                                        if publish_time == "未知":
                                            date_selectors = [
                                                '.date', '.time', '.published', '.timestamp',
                                                'time', '[itemprop="datePublished"]',
                                                '.article-date', '.post-date', '.entry-date',
                                                '.post-meta', '.article-meta'
                                            ]
                                            
                                            for selector in date_selectors:
                                                date_element = detail_soup.select_one(selector)
                                                if date_element:
                                                    publish_time = date_element.get_text(strip=True)
                                                    # 尝试清理日期文本
                                                    publish_time = re.sub(r'(Published|Updated|Posted|Date):?\s*', '', publish_time)
                                                    break
                                                    
                                    print(f"  发布时间: {publish_time}")
                                    
                                    # 将结果添加到Excel
                                    existing_links.add(link)
                                    result_sheet.append([source, link, title, publish_time])
                                    links_found = True
                                    
                                except Exception as e:
                                    print(f"❌ 获取详情页时出错: {str(e)}")
                                    # 即使出错也添加到结果，但没有发布时间
                                    existing_links.add(link)
                                    result_sheet.append([source, link, title, "未知"])
                                    links_found = True
                            else:
                                print(f"  跳过已存在的链接")
                    
                    print(f"共找到 {valid_links_count} 个有效新闻链接")
                    
                    if not links_found:
                        print("❌ 未找到任何有效新闻链接")
                        
                except Exception as e:
                    print(f"❌ 处理HTML时出错: {str(e)}")
            else:
                print(f"❌ 爬取失败: {result.error_message}")
    except Exception as e:
        print(f"❌ 处理异常: {str(e)}")
    
    # 立即保存结果，以防中途崩溃
    result_wb.save(result_file)
    print(f"保存当前结果到 {result_file}")

# 主程序
async def main():
    tasks = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        remark, main_url, source = row  # 更新解包以适应三列格式
        tasks.append(fetch_news_links(main_url, source))
    
    # 串行执行，避免并发问题
    for task in tasks:
        await task
    
    result_wb.save(result_file)
    print(f'所有任务完成，结果已保存到 {result_file}')

if __name__ == '__main__':
    asyncio.run(main()) 