import requests
import openpyxl
from datetime import datetime
import os
import sys
import asyncio
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig
from crawl4ai.content_filter_strategy import PruningContentFilter
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
from bs4 import BeautifulSoup

# 配置选项
SAVE_TO_EXCEL = True  # 设置为False可以禁用Excel输出功能

# 从命令行参数获取爬取范围，否则使用默认值
if len(sys.argv) >= 3:
    start_index = int(sys.argv[1])
    end_index = int(sys.argv[2])
else:

    start_index = 1
    end_index = 17

# 指定要测试的链接索引列表
indices = list(range(start_index, end_index + 1))
print(f"将爬取索引范围：{start_index} 到 {end_index}")

# 是否使用 ds 接口
use_ds_interface = False  # 设置为 True 使用 ds 接口 （暂定，现在还不用）

# 创建Excel文件保存结果
def create_result_excel():
    if not SAVE_TO_EXCEL:
        return None, None, None
        
    # 使用时间戳命名文件，避免覆盖
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = f'crawl_results_{timestamp}.xlsx'
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Crawl Results'
    
    # 设置标题行
    headers = ['索引', '原始链接', '标题', '正文内容', '提取方法', '字符数', '爬取时间', '状态', 'content_origin', 'after_prune']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header
    
    wb.save(result_file)
    print(f"创建结果文件：{result_file}")
    return wb, sheet, result_file

# 创建Excel结果文件
wb, sheet, result_file = create_result_excel()

# 计数器，用于记录当前行
row_counter = 2  # 从第2行开始，第1行是标题

# 读取原始链接
def get_original_links():
    links = {}
    try:
        source_wb = openpyxl.load_workbook('..\\testsample.xlsx')
        source_sheet = source_wb.active
        
        for i in indices:
            link = source_sheet.cell(row=i+1, column=1).value  # Excel行从1开始，跳过标题行
            if link:
                links[i] = link
                
    except Exception as e:
        print(f"读取原始链接出错: {str(e)}")
        
    return links

# 获取原始链接
original_links = get_original_links()

# 输出结果到Excel
def save_result_to_excel(index, link, title, content, extraction_method, char_count, crawl_time, status, content_origin, after_prune):
    if not SAVE_TO_EXCEL:
        return
        
    global row_counter
    
    # 填充结果
    sheet.cell(row=row_counter, column=1).value = index
    sheet.cell(row=row_counter, column=2).value = link
    sheet.cell(row=row_counter, column=3).value = title
    
    # 正文内容可能很长，截取一部分
    if content and len(content) > 32767:  # Excel单元格字符限制
        sheet.cell(row=row_counter, column=4).value = content[:32000] + "...(内容过长被截断)"
    else:
        sheet.cell(row=row_counter, column=4).value = content
    
    # 添加提取方法
    sheet.cell(row=row_counter, column=5).value = extraction_method
        
    sheet.cell(row=row_counter, column=6).value = char_count
    sheet.cell(row=row_counter, column=7).value = crawl_time
    sheet.cell(row=row_counter, column=8).value = status
    sheet.cell(row=row_counter, column=9).value = content_origin
    sheet.cell(row=row_counter, column=10).value = after_prune
    
    # 保存工作簿
    wb.save(result_file)
    
    # 增加行计数器
    row_counter += 1

# 记录总体统计信息
success_count = 0
error_count = 0

# 创建 PruningContentFilter 实例
prune_filter = PruningContentFilter(
    threshold=0.45,           # 调整阈值以保留更多内容
    threshold_type="dynamic",  # 使用动态阈值
    min_word_threshold=5      # 忽略少于5个词的节点
)

# 创建 Markdown Generator
md_generator = DefaultMarkdownGenerator(content_filter=prune_filter)

# 配置 CrawlerRunConfig
config = CrawlerRunConfig(markdown_generator=md_generator)

# 修改爬虫逻辑
for index in indices:
    print(f"爬取索引 {index}...")
    
    try:
        # 检查是否有此索引的链接
        if index not in original_links:
            print(f"索引 {index} 没有对应的链接")
            save_result_to_excel(index, "链接不存在", "", "", "无", 0, "", "错误：链接不存在", "", "")
            error_count += 1
            continue
            
        link = original_links[index]
        start_time = datetime.now()
        
        print(f"发送请求到 /crawl 端点，链接: {link}")
        # 使用异步爬虫
        async def fetch_content():
            async with AsyncWebCrawler() as crawler:
                result = await crawler.arun(url=link, config=config)
                return result
        
        result = asyncio.run(fetch_content())
        end_time = datetime.now()
        crawl_duration = (end_time - start_time).total_seconds()
        
        # 处理响应
        if result.success:
            # 使用 BeautifulSoup 提取标题
            soup = BeautifulSoup(result.html, 'html.parser')
            title = soup.title.string if soup.title else '未找到标题'
            content_origin = result.markdown.raw_markdown if hasattr(result.markdown, 'raw_markdown') else ''
            after_prune = result.markdown.fit_markdown if hasattr(result.markdown, 'fit_markdown') else ''
            content = result.markdown.fit_markdown
            extraction_method = "fit_markdown"
            char_count = len(content)
            
            print(f"成功爬取索引 {index}, 字符数: {char_count}, 提取方法: {extraction_method}")
            save_result_to_excel(
                index, 
                link, 
                title,
                content, 
                extraction_method,
                char_count, 
                f"{crawl_duration} 秒", 
                "成功",
                content_origin,
                after_prune
            )
            success_count += 1
        else:
            error_msg = result.error_message
            print(f"索引 {index} 爬取失败: {error_msg}")
            save_result_to_excel(
                index, 
                link, 
                "",
                "",
                "无",
                0, 
                f"{crawl_duration} 秒", 
                f"错误: {error_msg}",
                "",
                ""
            )
            error_count += 1
            
    except requests.exceptions.Timeout:
        print(f"索引 {index} 爬取超时")
        save_result_to_excel(
            index, 
            original_links.get(index, "未知链接"), 
            "",
            "",
            "超时",
            0, 
            "", 
            "错误: 请求超时",
            "",
            ""
        )
        error_count += 1
        
    except requests.exceptions.RequestException as e:
        print(f"索引 {index} 请求异常: {str(e)}")
        save_result_to_excel(
            index, 
            original_links.get(index, "未知链接"), 
            "",
            "",
            "请求异常",
            0, 
            "", 
            f"错误: 请求异常 - {str(e)}",
            "",
            ""
        )
        error_count += 1
        
    except Exception as e:
        print(f"索引 {index} 未知错误: {str(e)}")
        save_result_to_excel(
            index, 
            original_links.get(index, "未知链接"), 
            "",
            "",
            "未知错误",
            0, 
            "", 
            f"错误: 未知异常 - {str(e)}",
            "",
            ""
        )
        error_count += 1
        
    finally:
        print(f"完成索引 {index} 处理")
        print("-" * 40)

# 添加汇总行
if SAVE_TO_EXCEL:
    sheet.cell(row=row_counter, column=1).value = "汇总"
    sheet.cell(row=row_counter, column=3).value = f"总计爬取: {len(indices)}个链接, 成功: {success_count}, 失败: {error_count}"
    wb.save(result_file)
    print(f"已完成所有爬取任务，结果保存在 {result_file}")
else:
    print(f"爬取任务完成，Excel输出功能已禁用")

print(f"爬取任务统计: 总计尝试爬取 {len(indices)} 个链接, 成功: {success_count}, 失败: {error_count}")