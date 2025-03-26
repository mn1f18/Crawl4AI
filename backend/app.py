from flask import Flask, request, jsonify, render_template
import openpyxl
import asyncio
from crawl4ai import AsyncWebCrawler
from bs4 import BeautifulSoup
from datetime import datetime
import requests
from config import DS_API_KEY

app = Flask(__name__)

@app.route('/crawl', methods=['POST'])
def crawl():
    index = request.json.get('index', 1)  # 默认测试第一个链接
    use_ds = request.json.get('use_ds', False)  # 是否使用 ds 接口
    content_type = request.json.get('content_type', 'content')  # 获取内容类型，默认为'content'
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook('..\\testsample.xlsx')
        sheet = wb.active
        
        # 获取指定索引的链接
        link = sheet.cell(row=index+1, column=1).value  # Excel行从1开始，跳过标题行
        
        if not link:
            return jsonify({
                'url': f"索引 {index} 对应的链接不存在",
                'content': f"错误：索引 {index} 没有对应的链接",
                'duration': 0,
                'index': index,
                'error': True
            })
            
        # 爬取指定链接
        try:
            result = asyncio.run(crawl_link(link, content_type, index))
        except Exception as e:
            return jsonify({
                'url': link,
                'content': f"爬取错误：{str(e)}",
                'duration': 0,
                'index': index,
                'error': True
            })
        
        # 处理 ds 接口
        ds_result = 'DS interface not used'
        if use_ds:
            try:
                ds_result = process_with_ds(result['content'])
                result['ds_result'] = ds_result
            except Exception as e:
                result['ds_result'] = f"DS接口错误：{str(e)}"
        
        # 返回结果，但不渲染HTML
        return jsonify({
            'url': result['url'],
            'content': result['content'],
            'duration': result['duration'],
            'index': result['index'],
            'error': False
        })
        
    except Exception as e:
        # 处理所有其他异常
        return jsonify({
            'url': f"索引 {index}",
            'content': f"处理错误：{str(e)}",
            'duration': 0,
            'index': index,
            'error': True
        })

@app.route('/view_result', methods=['GET'])
def view_result():
    content_type = request.args.get('type', 'content')  # 获取查询参数，默认为'content'
    
    # 获取最大链接数，默认为10
    max_links = int(request.args.get('max', 10))
    
    # 获取自定义索引列表（如果有）
    custom_indices = request.args.get('indices', None)
    indices_to_crawl = []
    
    if custom_indices:
        # 支持格式如 "1,2,5-8,10"
        try:
            parts = custom_indices.split(',')
            for part in parts:
                if '-' in part:
                    # 处理范围，如 "5-8"
                    start, end = map(int, part.split('-'))
                    indices_to_crawl.extend(range(start, end + 1))
                else:
                    # 处理单个索引，如 "1"
                    indices_to_crawl.append(int(part))
        except ValueError:
            # 如果格式错误，使用默认范围
            indices_to_crawl = list(range(1, max_links + 1))
    else:
        # 使用默认范围
        indices_to_crawl = list(range(1, max_links + 1))
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook('..\\testsample.xlsx')
        sheet = wb.active
        
        # 获取链接
        links = []
        for i in indices_to_crawl:
            link = sheet.cell(row=i+1, column=1).value  # Excel行从1开始，跳过标题行
            if link:
                links.append((i, link))  # 保存索引和链接
        
        # 并行爬取所有链接
        results = []
        for idx, link in links:
            try:
                result = asyncio.run(crawl_link(link, content_type, idx))
                results.append(result)
            except Exception as e:
                # 如果爬取过程中出错，添加错误信息到结果列表
                error_result = {
                    'url': link,
                    'content': f"爬取错误: {str(e)}",
                    'duration': 0,
                    'index': idx,
                    'error': True
                }
                results.append(error_result)
        
        # 渲染HTML模板
        return render_template('results.html', results=results, type=content_type)
    
    except Exception as e:
        # 处理整体异常
        error_result = [{
            'url': '加载错误',
            'content': f"页面加载错误: {str(e)}",
            'duration': 0,
            'index': 0,
            'error': True
        }]
        return render_template('results.html', results=error_result, type=content_type)

@app.route('/view_truecontent', methods=['GET'])
def view_truecontent():
    # 获取正文内容
    content_type = 'truecontent'
    
    # 获取最大链接数，默认为10
    max_links = int(request.args.get('max', 10))
    
    # 获取自定义索引列表（如果有）
    custom_indices = request.args.get('indices', None)
    indices_to_crawl = []
    
    if custom_indices:
        # 支持格式如 "1,2,5-8,10"
        try:
            parts = custom_indices.split(',')
            for part in parts:
                if '-' in part:
                    # 处理范围，如 "5-8"
                    start, end = map(int, part.split('-'))
                    indices_to_crawl.extend(range(start, end + 1))
                else:
                    # 处理单个索引，如 "1"
                    indices_to_crawl.append(int(part))
        except ValueError:
            # 如果格式错误，使用默认范围
            indices_to_crawl = list(range(1, max_links + 1))
    else:
        # 使用默认范围
        indices_to_crawl = list(range(1, max_links + 1))
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook('..\\testsample.xlsx')
        sheet = wb.active
        
        # 获取链接
        links = []
        for i in indices_to_crawl:
            link = sheet.cell(row=i+1, column=1).value  # Excel行从1开始，跳过标题行
            if link:
                links.append((i, link))  # 保存索引和链接
        
        # 爬取所有链接
        results = []
        for idx, link in links:
            try:
                result = asyncio.run(crawl_link(link, content_type, idx))
                results.append(result)
            except Exception as e:
                # 如果爬取过程中出错，添加错误信息到结果列表
                error_result = {
                    'url': link,
                    'content': f"爬取错误: {str(e)}",
                    'duration': 0,
                    'index': idx,
                    'error': True,
                    'token_count': 0
                }
                results.append(error_result)
        
        # 计算字符数量并处理DS接口
        use_ds = request.args.get('use_ds', 'false').lower() == 'true'
        
        for result in results:
            # 如果是错误结果，跳过DS处理
            if result.get('error', False):
                result['ds_result'] = '爬取出错，无法处理DS接口'
                continue
                
            # 计算字符数
            result['token_count'] = len(result['content'])
            
            # 处理DS接口
            result['ds_result'] = 'DS接口未使用'
            if use_ds:
                try:
                    result['ds_result'] = process_with_ds(result['content'])
                except Exception as e:
                    result['ds_result'] = f"DS接口错误: {str(e)}"
        
        # 渲染HTML模板
        return render_template('truecontent_multi.html', results=results, prompt='default_prompt')
        
    except Exception as e:
        # 处理整体异常
        error_result = [{
            'url': '加载错误',
            'content': f"页面加载错误: {str(e)}",
            'duration': 0,
            'index': 0,
            'error': True,
            'token_count': 0,
            'ds_result': '页面错误，无法处理'
        }]
        return render_template('truecontent_multi.html', results=error_result, prompt='default_prompt')

@app.route('/ds_check', methods=['POST'])
def ds_check():
    # 从请求中获取正文内容
    content = request.json.get('content', '')
    
    # 使用 ds 接口进行处理
    # 这里假设有一个函数 `process_with_ds` 来处理内容
    result = process_with_ds(content)
    
    # 返回处理结果
    return jsonify({'result': result})

# 假设的处理函数
# 这里你需要实现与 ds 接口的实际交互逻辑
def process_with_ds(content):
    # 使用 DS_API_KEY 进行处理
    url = "https://api.ds.example.com/analyze"  # 假设的 ds 接口 URL
    headers = {
        "Authorization": f"Bearer {DS_API_KEY}",
        "Content-Type": "application/json"
    }
    data = {
        "content": content
    }
    response = requests.post(url, headers=headers, json=data)
    if response.status_code == 200:
        return response.json().get('analysis', 'No analysis found')
    else:
        return f"Error: {response.status_code}"

async def crawl_links(links, content_type='article'):
    async with AsyncWebCrawler() as crawler:
        tasks = [crawl_link(link, content_type, index) for index, link in enumerate(links, start=1)]
        return await asyncio.gather(*tasks)

async def crawl_link(link, content_type='article', index=None):
    try:
        async with AsyncWebCrawler() as crawler:
            start_time = datetime.now()
            try:
                # 对于truecontent类型，使用高级内容过滤器
                if content_type == 'truecontent':
                    try:
                        # 添加版本信息调试输出
                        import inspect
                        import crawl4ai
                        print(f"\n===== 调试信息 =====")
                        print(f"Crawl4AI版本: {getattr(crawl4ai, '__version__', '未知')}")
                        print(f"爬取链接: {link}")
                        
                        # 导入必要的类
                        from crawl4ai.content_filter_strategy import PruningContentFilter, BM25ContentFilter
                        from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
                        from crawl4ai.async_configs import CrawlerRunConfig
                        
                        # 检查导入的类是否存在预期方法
                        print(f"PruningContentFilter是否可用: {inspect.isclass(PruningContentFilter)}")
                        print(f"DefaultMarkdownGenerator是否可用: {inspect.isclass(DefaultMarkdownGenerator)}")
                        
                        # 创建 PruningContentFilter 实例
                        prune_filter = PruningContentFilter(
                            threshold=0.35,           # 调低阈值以保留更多内容
                            threshold_type="dynamic", # 动态阈值更灵活
                            min_word_threshold=3      # 降低词数阈值让更多内容通过
                        )
                        print(f"PruningContentFilter实例创建成功: {prune_filter is not None}")
                        
                        # 创建markdown生成器并应用过滤器
                        md_generator = DefaultMarkdownGenerator(content_filter=prune_filter)
                        print(f"DefaultMarkdownGenerator实例创建成功: {md_generator is not None}")
                        
                        # 创建爬虫配置
                        config = CrawlerRunConfig(
                            markdown_generator=md_generator,
                            word_count_threshold=5,                # 降低词数阈值
                            excluded_tags=["nav", "footer", "header"], # 排除这些标签
                            exclude_external_links=True,            # 排除外部链接
                            remove_overlay_elements=True            # 移除弹窗覆盖元素
                        )
                        print(f"CrawlerRunConfig实例创建成功: {config is not None}")
                        print(f"配置使用的markdown生成器: {config.markdown_generator}")
                        
                        # 使用配置运行爬虫
                        print(f"开始执行爬取...")
                        result = await crawler.arun(url=link, config=config)
                        print(f"爬取完成，检查返回结果...")
                        
                        # 检查结果对象
                        print(f"Result对象类型: {type(result)}")
                        print(f"Result.markdown类型: {type(getattr(result, 'markdown', None))}")
                        print(f"Result.markdown是否具有fit_markdown属性: {hasattr(getattr(result, 'markdown', None), 'fit_markdown')}")
                        
                        if hasattr(result, 'markdown') and hasattr(result.markdown, 'fit_markdown'):
                            print(f"fit_markdown是否为空: {not bool(result.markdown.fit_markdown)}")
                            if result.markdown.fit_markdown:
                                print(f"fit_markdown前50个字符: {result.markdown.fit_markdown[:50]}")
                            else:
                                print("fit_markdown为空")
                        else:
                            print("markdown对象缺少fit_markdown属性")
                        
                        # 检查raw_markdown是否存在
                        if hasattr(result, 'markdown') and hasattr(result.markdown, 'raw_markdown'):
                            print(f"raw_markdown是否存在: 是，长度为{len(result.markdown.raw_markdown)}")
                        else:
                            print("raw_markdown属性不存在")
                        
                        print(f"===== 调试信息结束 =====\n")
                        
                    except (ImportError, AttributeError) as e:
                        # 如果导入失败或不支持这些类，使用基本配置
                        print(f"高级内容过滤不可用，将使用基本配置: {str(e)}")
                        print(f"错误类型: {type(e).__name__}")
                        print(f"错误位置: {getattr(e, '__traceback__', '未知')}")
                        result = await crawler.arun(url=link)
                elif content_type == 'article':
                    # 使用article方法提取，利用选择器精准识别文章正文
                    print(f"使用article方法提取链接: {link}")
                    
                    # 文章内容选择器
                    article_selectors = [
                        "article", "main", ".article", ".post", ".content", 
                        ".article-content", ".post-content", ".entry-content",
                        "#article", "#content", "#main-content", ".main-content",
                        "[itemprop='articleBody']", ".story-body", ".news-content"
                    ]
                    
                    # 创建文章提取配置
                    from crawl4ai.async_configs import CrawlerRunConfig
                    
                    # 首先尝试基础配置
                    result = await crawler.arun(url=link)
                    
                    # 如果获取成功，尝试使用选择器提取
                    if result and result.html:
                        article_content = None
                        content_source = "默认提取"
                        
                        # 尝试各个选择器
                        for selector in article_selectors:
                            try:
                                article_config = CrawlerRunConfig(
                                    css_selector=selector,
                                    excluded_tags=["nav", "header", "footer", "aside", "form"],
                                    exclude_external_links=True,
                                    word_count_threshold=5
                                )
                                
                                article_result = await crawler.arun(url=link, config=article_config)
                                
                                # 检查提取的内容
                                if article_result and hasattr(article_result, 'markdown'):
                                    article_markdown = None
                                    if isinstance(article_result.markdown, str):
                                        article_markdown = article_result.markdown
                                    elif hasattr(article_result.markdown, 'raw_markdown') and article_result.markdown.raw_markdown:
                                        article_markdown = article_result.markdown.raw_markdown
                                        
                                    # 如果内容非空且长度合适，使用这个选择器的结果
                                    if article_markdown and len(article_markdown) > 100:
                                        # 更新result以使用文章选择器提取的结果
                                        result = article_result
                                        content_source = f"选择器 '{selector}'"
                                        print(f"成功使用文章选择器 '{selector}' 提取内容")
                                        break
                            except Exception as e:
                                print(f"选择器 '{selector}' 提取失败: {str(e)}")
                                continue
                else:
                    # 对其他类型使用默认配置
                    result = await crawler.arun(url=link)
                
                end_time = datetime.now()
                crawl_duration = (end_time - start_time).total_seconds()
            except Exception as e:
                print(f"爬取过程中出现错误: {str(e)}")
                print(f"错误类型: {type(e).__name__}")
                import traceback
                print(f"错误堆栈: {traceback.format_exc()}")
                return {
                    'url': link,
                    'content': f"爬取失败: {str(e)}",
                    'duration': (datetime.now() - start_time).total_seconds(),
                    'index': index,
                    'error': True
                }

            try:
                if content_type == 'title':
                    # 从 HTML 中提取标题
                    soup = BeautifulSoup(result.html, 'html.parser')
                    title = soup.title.string if soup.title else 'No title found'
                    return {
                        'url': link,
                        'content': title,
                        'duration': crawl_duration,
                        'index': index,
                        'error': False
                    }
                elif content_type == 'link':
                    # 提取链接字符串
                    all_links = [link['href'] for link in result.links.get('internal', []) + result.links.get('external', []) if 'href' in link]
                    return {
                        'url': link,
                        'content': '\n'.join(all_links),
                        'duration': crawl_duration,
                        'index': index,
                        'error': False
                    }
                elif content_type == 'truecontent':
                    # 使用改进的内容提取逻辑，优先顺序：
                    # 1. 使用Crawl4AI的fit_markdown（如果可用）
                    # 2. 使用原始markdown
                    # 3. 使用选择器提取正文
                    # 4. 最后尝试获取所有文本内容
                    
                    content = None
                    content_source = "未知"
                    
                    # 1. 首先尝试使用fit_markdown
                    if (hasattr(result, 'markdown') and 
                        not isinstance(result.markdown, str) and 
                        hasattr(result.markdown, 'fit_markdown') and 
                        result.markdown.fit_markdown):
                        content = result.markdown.fit_markdown
                        content_source = "fit_markdown智能提取"
                        print(f"成功使用fit_markdown提取内容，长度: {len(content)}")
                    else:
                        print(f"fit_markdown提取失败，尝试其他方法")
                        if hasattr(result, 'markdown'):
                            print(f"markdown类型: {type(result.markdown)}")
                            print(f"是否有fit_markdown属性: {hasattr(result.markdown, 'fit_markdown')}")
                            if hasattr(result.markdown, 'fit_markdown'):
                                print(f"fit_markdown是否为空: {not bool(result.markdown.fit_markdown)}")
                        else:
                            print("result没有markdown属性")
                    
                    # 2. 然后尝试使用原始markdown
                    if not content or content.strip() == '':
                        if hasattr(result, 'markdown'):
                            # 检查是否有原始markdown
                            if isinstance(result.markdown, str):
                                content = result.markdown
                                content_source = "原始markdown"
                                print(f"使用原始markdown字符串，长度: {len(content)}")
                            elif hasattr(result.markdown, 'raw_markdown') and result.markdown.raw_markdown:
                                content = result.markdown.raw_markdown
                                content_source = "原始markdown"
                                print(f"使用raw_markdown属性，长度: {len(content)}")
                    
                    # 3. 如果markdown失败，尝试使用CSS选择器
                    if not content or content.strip() == '':
                        print(f"尝试使用CSS选择器提取内容")
                        soup = BeautifulSoup(result.html, 'html.parser')
                        
                        # 多个可能的内容选择器，按优先级排序
                        content_selectors = [
                            'article', 'main', 
                            'div.content', 'div.article-content', 
                            'div.post-content', 'div.entry-content',
                            '.article-body', '.news-content', '#content',
                            '.story-body', '.entry', '.post',
                            'section.content', '[itemprop="articleBody"]',
                            '.article', '.blog-post', '.main-content'
                        ]
                        
                        # 尝试每个选择器
                        for selector in content_selectors:
                            content_element = soup.select_one(selector)
                            if content_element:
                                # 移除可能影响内容的元素
                                for el in content_element.select('script, style, nav, header, footer, aside, .ad, .ads, .advertisement, .sidebar'):
                                    el.decompose()
                                    
                                content = content_element.get_text(strip=True)
                                content_source = f"选择器 '{selector}'"
                                print(f"成功使用选择器 '{selector}' 提取内容，长度: {len(content)}")
                                break
                        
                        if not content or content.strip() == '':
                            print(f"所有选择器都失败")
                    
                    # 4. 如果选择器也失败，获取所有文本
                    if not content or content.strip() == '':
                        print(f"尝试提取全文文本")
                        soup = BeautifulSoup(result.html, 'html.parser')
                        # 移除不需要的元素
                        for tag in ['script', 'style', 'nav', 'header', 'footer', 'aside']:
                            for element in soup.find_all(tag):
                                element.decompose()
                        content = soup.get_text(strip=True)
                        content_source = "全文文本"
                        print(f"使用全文提取，长度: {len(content)}")
                            
                    # 如果内容仍然为空，返回错误消息
                    if not content or content.strip() == '':
                        content = 'No content could be extracted from this page'
                        content_source = "无法提取"
                        print(f"所有提取方法都失败，无法获取内容")
                    
                    # 为内容添加来源信息
                    content = f"[内容来源: {content_source}]\n\n{content}"
                        
                    return {
                        'url': link,
                        'content': content,
                        'duration': crawl_duration,
                        'index': index,
                        'error': False,
                        'content_source': content_source
                    }
                elif content_type == 'article':
                    # article方法专注于提取干净的文章正文
                    content = None
                    content_source = "默认提取"
                    
                    # 尝试获取markdown内容
                    if hasattr(result, 'markdown'):
                        if isinstance(result.markdown, str):
                            content = result.markdown
                        elif hasattr(result.markdown, 'raw_markdown') and result.markdown.raw_markdown:
                            content = result.markdown.raw_markdown
                    
                    # 如果是通过选择器提取的内容，添加来源信息
                    if 'content_source' in locals() and content_source != "默认提取":
                        content = f"[内容来源: {content_source}]\n\n{content}"
                    
                    return {
                        'url': link,
                        'content': content,
                        'duration': crawl_duration,
                        'index': index,
                        'error': False,
                        'content_source': content_source if 'content_source' in locals() else "智能提取"
                    }
                else:
                    # 默认使用article方法作为兜底，防止旧代码调用content方法时返回空内容
                    content = None
                    if hasattr(result, 'markdown'):
                        if isinstance(result.markdown, str):
                            content = result.markdown
                        elif hasattr(result.markdown, 'raw_markdown') and result.markdown.raw_markdown:
                            content = result.markdown.raw_markdown
                    else:
                        content = 'No content'
                        
                    return {
                        'url': link,
                        'content': content,
                        'duration': crawl_duration,
                        'index': index,
                        'error': False
                    }
            except Exception as e:
                print(f"内容处理过程中出现错误: {str(e)}")
                print(f"错误类型: {type(e).__name__}")
                import traceback
                print(f"处理错误堆栈: {traceback.format_exc()}")
                return {
                    'url': link,
                    'content': f"内容处理失败: {str(e)}",
                    'duration': crawl_duration,
                    'index': index,
                    'error': True
                }
    except Exception as e:
        print(f"爬虫初始化失败: {str(e)}")
        print(f"错误类型: {type(e).__name__}")
        import traceback
        print(f"初始化错误堆栈: {traceback.format_exc()}")
        return {
            'url': link,
            'content': f"爬虫初始化失败: {str(e)}",
            'duration': 0,
            'index': index,
            'error': True
        }

@app.route('/view_all_results', methods=['GET'])
def view_all_results():
    content_type = request.args.get('type', 'content')  # 获取查询参数，默认为'content'
    
    # 获取最大链接数，默认为10
    max_links = int(request.args.get('max', 10))
    
    # 获取自定义索引列表（如果有）
    custom_indices = request.args.get('indices', None)
    indices_to_crawl = []
    
    if custom_indices:
        # 支持格式如 "1,2,5-8,10"
        try:
            parts = custom_indices.split(',')
            for part in parts:
                if '-' in part:
                    # 处理范围，如 "5-8"
                    start, end = map(int, part.split('-'))
                    indices_to_crawl.extend(range(start, end + 1))
                else:
                    # 处理单个索引，如 "1"
                    indices_to_crawl.append(int(part))
        except ValueError:
            # 如果格式错误，使用默认范围
            indices_to_crawl = list(range(1, max_links + 1))
    else:
        # 使用默认范围
        indices_to_crawl = list(range(1, max_links + 1))
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook('..\\testsample.xlsx')
        sheet = wb.active
        
        # 获取链接
        links = []
        for i in indices_to_crawl:
            link = sheet.cell(row=i+1, column=1).value  # Excel行从1开始，跳过标题行
            if link:
                links.append((i, link))  # 保存索引和链接
        
        # 爬取所有链接
        results = []
        for idx, link in links:
            try:
                result = asyncio.run(crawl_link(link, content_type, idx))
                results.append(result)
            except Exception as e:
                # 如果爬取过程中出错，添加错误信息到结果列表
                error_result = {
                    'url': link,
                    'content': f"爬取错误: {str(e)}",
                    'duration': 0,
                    'index': idx,
                    'error': True
                }
                results.append(error_result)
        
        # 渲染HTML模板
        return render_template('all_results.html', results=results)
        
    except Exception as e:
        # 处理整体异常
        error_result = [{
            'url': '加载错误',
            'content': f"页面加载错误: {str(e)}",
            'duration': 0,
            'index': 0,
            'error': True
        }]
        return render_template('all_results.html', results=error_result)

@app.route('/extract_content', methods=['POST'])
def extract_content():
    """
    专门用于测试自动识别正文的接口
    请求参数: 
        {
            "url": "要爬取的URL",
            "index": 可选的索引编号
        }
    """
    try:
        # 获取请求参数
        data = request.json
        url = data.get('url')
        index = data.get('index', None)
        
        if not url:
            return jsonify({
                'success': False,
                'error': '未提供URL',
                'content': '',
                'extraction_method': '无',
                'index': index
            })
        
        # 调用爬虫提取正文
        print(f"尝试从URL提取正文: {url}")
        try:
            # 使用AsyncWebCrawler和改进的CrawlerRunConfig
            start_time = datetime.now()
            
            async def smart_extract_content():
                from crawl4ai import AsyncWebCrawler, CrawlerRunConfig
                
                # 可能的文章内容选择器
                article_selectors = [
                    "article", "main", ".article", ".post", ".content", 
                    ".article-content", ".post-content", ".entry-content",
                    "#article", "#content", "#main-content", ".main-content",
                    "[itemprop='articleBody']", ".story-body", ".news-content"
                ]
                
                # 可能的标题选择器
                title_selectors = [
                    "h1", "h1.title", "h1.article-title", "h1.post-title", 
                    ".article-title", ".post-title", ".entry-title",
                    "[itemprop='headline']", ".headline", ".title"
                ]
                
                # 创建初始基本配置
                base_config = CrawlerRunConfig(
                    # 内容过滤基本设置
                    word_count_threshold=5,  # 最小文本块字数
                    excluded_tags=["nav", "header", "footer", "aside", "form"],
                    exclude_external_links=True,
                    exclude_social_media_links=True,
                    remove_overlay_elements=True  # 移除覆盖元素
                )
                
                extraction_result = {
                    'title': None,
                    'content': None,
                    'extraction_method': '基础提取',
                    'article_selector_used': None,
                    'title_selector_used': None
                }
                
                try:
                    # 开始爬取
                    async with AsyncWebCrawler() as crawler:
                        # 首先使用基本配置爬取
                        result = await crawler.arun(url=url, config=base_config)
                        
                        if not result or not result.html:
                            return {
                                'success': False,
                                'error': '无法获取网页内容',
                                'extraction_method': '爬取失败'
                            }
                        
                        # 尝试从HTML中提取标题
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(result.html, 'html.parser')
                        
                        # 首先使用页面标题作为后备
                        if soup and soup.title:
                            extraction_result['title'] = soup.title.string
                        
                        # 尝试各种标题选择器
                        for selector in title_selectors:
                            try:
                                title_element = soup.select_one(selector)
                                if title_element and title_element.get_text().strip():
                                    extraction_result['title'] = title_element.get_text().strip()
                                    extraction_result['title_selector_used'] = selector
                                    break
                            except Exception as e:
                                print(f"标题选择器 {selector} 失败: {str(e)}")
                                continue
                        
                        # 记录原始提取的markdown内容
                        original_markdown = ''
                        if hasattr(result, 'markdown'):
                            if isinstance(result.markdown, str):
                                original_markdown = result.markdown
                            elif hasattr(result.markdown, 'raw_markdown') and result.markdown.raw_markdown:
                                original_markdown = result.markdown.raw_markdown
                        
                        # 尝试不同的文章选择器
                        article_content = None
                        for selector in article_selectors:
                            try:
                                # 为每个选择器创建新的配置
                                article_config = CrawlerRunConfig(
                                    css_selector=selector,  # 使用CSS选择器
                                    excluded_tags=["nav", "header", "footer", "aside", "form"],
                                    exclude_external_links=True,
                                    word_count_threshold=5
                                )
                                
                                article_result = await crawler.arun(url=url, config=article_config)
                                
                                if not article_result:
                                    print(f"选择器 {selector} 返回无效结果")
                                    continue
                                
                                # 检查提取的内容
                                article_markdown = None
                                if hasattr(article_result, 'markdown'):
                                    if isinstance(article_result.markdown, str):
                                        article_markdown = article_result.markdown
                                    elif hasattr(article_result.markdown, 'raw_markdown') and article_result.markdown.raw_markdown:
                                        article_markdown = article_result.markdown.raw_markdown
                                        
                                # 如果内容非空且长度合适，使用这个选择器的结果
                                if article_markdown and len(article_markdown) > 100:
                                    article_content = article_markdown
                                    extraction_result['article_selector_used'] = selector
                                    extraction_result['extraction_method'] = f'CSS选择器({selector})'
                                    break
                            except Exception as e:
                                print(f"选择器 {selector} 提取失败: {str(e)}")
                                continue
                        
                        # 如果前面的选择器都失败了，尝试使用target_elements
                        if not article_content:
                            try:
                                # 使用target_elements参数
                                target_config = CrawlerRunConfig(
                                    target_elements=article_selectors[:5],  # 使用前5个选择器作为目标元素
                                    excluded_tags=["nav", "header", "footer", "aside", "form"],
                                    exclude_external_links=True,
                                    word_count_threshold=5
                                )
                                
                                target_result = await crawler.arun(url=url, config=target_config)
                                
                                if target_result and hasattr(target_result, 'markdown'):
                                    if isinstance(target_result.markdown, str):
                                        article_content = target_result.markdown
                                    elif hasattr(target_result.markdown, 'raw_markdown') and target_result.markdown.raw_markdown:
                                        article_content = target_result.markdown.raw_markdown
                                    
                                    if article_content and len(article_content) > 100:
                                        extraction_result['extraction_method'] = 'target_elements提取'
                            except Exception as e:
                                print(f"target_elements提取失败: {str(e)}")
                        
                        # 如果找到适合的内容，使用它
                        if article_content:
                            extraction_result['content'] = article_content
                        else:
                            # 否则使用原始提取的markdown
                            extraction_result['content'] = original_markdown
                            extraction_result['extraction_method'] = '全页面提取'
                        
                        # 最后的内容检查
                        if not extraction_result['content'] or len(extraction_result['content'].strip()) == 0:
                            # 尝试直接从HTML提取文本
                            try:
                                if soup:
                                    # 移除不需要的元素
                                    for tag in ['script', 'style', 'nav', 'header', 'footer', 'aside']:
                                        for element in soup.find_all(tag):
                                            element.decompose()
                                    
                                    text_content = soup.get_text(strip=True)
                                    if text_content and len(text_content) > 0:
                                        extraction_result['content'] = text_content
                                        extraction_result['extraction_method'] = 'HTML直接提取'
                            except Exception as e:
                                print(f"HTML直接提取失败: {str(e)}")
                        
                        # 如果还是没有内容，返回错误
                        if not extraction_result['content'] or len(extraction_result['content'].strip()) == 0:
                            extraction_result['content'] = "无法提取到有效内容"
                        
                        # 分析结果长度
                        extraction_result['char_count'] = len(extraction_result['content']) if extraction_result['content'] else 0
                        
                        return {
                            'success': True,
                            **extraction_result
                        }
                except Exception as e:
                    import traceback
                    print(f"爬取处理异常: {str(e)}\n{traceback.format_exc()}")
                    return {
                        'success': False,
                        'error': f"内部处理异常: {str(e)}",
                        'extraction_method': '内部错误'
                    }
            
            # 执行智能提取
            result = asyncio.run(smart_extract_content())
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            
            if not result.get('success', False):
                return jsonify({
                    'success': False,
                    'error': result.get('error', '内容提取失败'),
                    'content': '',
                    'extraction_method': result.get('extraction_method', '失败'),
                    'url': url,
                    'index': index
                })
            
            # 构建返回结果
            return jsonify({
                'success': True,
                'title': result.get('title', '未找到标题'),
                'content': result.get('content', ''),
                'extraction_method': result.get('extraction_method', '未知方法'),
                'title_selector': result.get('title_selector_used', '无'),
                'article_selector': result.get('article_selector_used', '无'),
                'url': url,
                'index': index,
                'char_count': result.get('char_count', 0),
                'duration': duration
            })
                
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"爬取URL时发生错误: {str(e)}\n{error_trace}")
            return jsonify({
                'success': False,
                'error': f"爬取错误: {str(e)}",
                'content': '',
                'extraction_method': '错误',
                'url': url,
                'index': index
            })
            
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"接口处理错误: {str(e)}\n{error_trace}")
        return jsonify({
            'success': False,
            'error': f"处理错误: {str(e)}",
            'content': '',
            'extraction_method': '错误',
            'index': index
        })

if __name__ == '__main__':
    app.run(debug=True)  
