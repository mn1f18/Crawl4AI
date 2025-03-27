import requests
import openpyxl
from datetime import datetime
import os
import sys
import traceback
from openpyxl.styles import Alignment, Font, PatternFill

# 配置选项
SAVE_TO_EXCEL = True  # 设置为False可以禁用Excel输出功能

# 从命令行参数获取爬取范围，否则使用默认值
if len(sys.argv) >= 3:
    start_index = int(sys.argv[1])
    end_index = int(sys.argv[2])
else:
    # 默认爬取范围
    start_index = 1
    end_index = 2

# 指定要测试的链接索引列表
indices = list(range(start_index, end_index + 1))
print(f"将爬取索引范围：{start_index} 到 {end_index}")

# 创建Excel文件保存结果
def create_result_excel():
    if not SAVE_TO_EXCEL:
        return None, None, None
        
    # 使用固定名称
    result_file = f'测试只爬取正文.xlsx'
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Crawl Results'
    
    # 设置标题行
    headers = ['索引', '原始链接', '文章标题', '正文内容', '提取方法', '标题选择器', '内容选择器', '字符数', '爬取时间', '状态', '重试次数']
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # 设置列宽
    sheet.column_dimensions['A'].width = 8   # 索引
    sheet.column_dimensions['B'].width = 40  # 原始链接
    sheet.column_dimensions['C'].width = 40  # 文章标题
    sheet.column_dimensions['D'].width = 60  # 正文内容
    sheet.column_dimensions['E'].width = 20  # 提取方法
    sheet.column_dimensions['F'].width = 20  # 标题选择器
    sheet.column_dimensions['G'].width = 20  # 内容选择器
    sheet.column_dimensions['H'].width = 12  # 字符数
    sheet.column_dimensions['I'].width = 15  # 爬取时间
    sheet.column_dimensions['J'].width = 25  # 状态
    sheet.column_dimensions['K'].width = 15  # 重试次数
    
    wb.save(result_file)
    print(f"创建结果文件：{result_file}")
    return wb, sheet, result_file

# 创建Excel结果文件
wb, sheet, result_file = create_result_excel()

# 计数器，用于记录当前行
row_counter = 2  # 从第2行开始，第1行是标题

# 读取原始链接
def get_original_links():
    links = {}
    try:
        source_wb = openpyxl.load_workbook('..\\testsample.xlsx')
        source_sheet = source_wb.active
        
        for i in indices:
            link = source_sheet.cell(row=i+1, column=1).value  # Excel行从1开始，跳过标题行
            if link:
                links[i] = link
                
    except Exception as e:
        print(f"读取原始链接出错: {str(e)}")
        
    return links

# 获取原始链接
original_links = get_original_links()

# 输出结果到Excel
def save_result_to_excel(index, link, title, content, extraction_method, title_selector, article_selector, char_count, crawl_time, status, retries):
    if not SAVE_TO_EXCEL:
        return
        
    global row_counter
    
    # 设置单元格样式
    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # 填充结果
    sheet.cell(row=row_counter, column=1).value = index
    sheet.cell(row=row_counter, column=2).value = link
    sheet.cell(row=row_counter, column=3).value = title
    
    # 设置所有单元格自动换行
    for col in range(1, 12):
        sheet.cell(row=row_counter, column=col).alignment = wrap_alignment
    
    # 正文内容可能很长，截取一部分
    if content and len(content) > 32767:  # Excel单元格字符限制
        sheet.cell(row=row_counter, column=4).value = content[:32000] + "...(内容过长被截断)"
    else:
        sheet.cell(row=row_counter, column=4).value = content
    
    # 添加提取方法和选择器信息
    sheet.cell(row=row_counter, column=5).value = extraction_method
    sheet.cell(row=row_counter, column=6).value = title_selector
    sheet.cell(row=row_counter, column=7).value = article_selector
        
    sheet.cell(row=row_counter, column=8).value = char_count
    sheet.cell(row=row_counter, column=9).value = crawl_time
    sheet.cell(row=row_counter, column=10).value = status
    sheet.cell(row=row_counter, column=11).value = retries
    
    # 根据爬取状态设置填充颜色
    if "成功" in status:
        for col in range(1, 12):
            sheet.cell(row=row_counter, column=col).fill = success_fill
    elif "错误" in status:
        for col in range(1, 12):
            sheet.cell(row=row_counter, column=col).fill = error_fill
    
    # 保存工作簿
    wb.save(result_file)
    
    # 增加行计数器
    row_counter += 1

# 错误重试机制
max_retries = 3

# 定期保存功能
save_interval = 100  # 每100次保存一次

# 记录总体统计信息
success_count = 0
error_count = 0
retry_count = 0

# 直接使用新的/extract_content接口获取正文
for index in indices:
    print(f"爬取索引 {index}...")
    retries = 0
    while retries <= max_retries:
        try:
            # 检查是否有此索引的链接
            if index not in original_links:
                print(f"索引 {index} 没有对应的链接")
                save_result_to_excel(index, "链接不存在", "", "", "无", "无", "无", 0, "", "错误：链接不存在", retries)
                error_count += 1
                break

            link = original_links[index]
            start_time = datetime.now()

            print(f"发送请求到 /extract_content 接口，链接: {link}")

            # 直接使用新接口提取正文
            response = requests.post(
                'http://127.0.0.1:5000/extract_content', 
                json={'url': link, 'index': index},
                timeout=180  # 设置更长的超时时间
            )

            end_time = datetime.now()
            crawl_duration = (end_time - start_time).total_seconds()

            # 处理响应
            if response.status_code == 200:
                result = response.json()
                if result.get('success', False):
                    title = result.get('title', '未找到标题')
                    content = result.get('content', '')
                    extraction_method = result.get('extraction_method', '未知方法')
                    title_selector = result.get('title_selector', '无')
                    article_selector = result.get('article_selector', '无')
                    char_count = result.get('char_count', len(content))

                    save_result_to_excel(
                        index, 
                        link,
                        title,
                        content, 
                        extraction_method,
                        title_selector,
                        article_selector,
                        char_count, 
                        f"{crawl_duration} 秒", 
                        "成功",
                        retries
                    )
                    success_count += 1
                    break
                else:
                    error_msg = result.get('error', '未知错误')
                    print(f"索引 {index} 爬取失败: {error_msg}")
                    retries += 1
                    if retries > max_retries:
                        save_result_to_excel(
                            index, 
                            link,
                            "错误", 
                            error_msg, 
                            "错误",
                            "无",
                            "无",
                            0, 
                            f"{crawl_duration} 秒", 
                            f"错误: {error_msg}",
                            retries
                        )
                        error_count += 1
            else:
                error_msg = f"HTTP错误: {response.status_code}"
                print(f"索引 {index} 爬取失败: {error_msg}")
                retries += 1
                if retries > max_retries:
                    save_result_to_excel(
                        index, 
                        link,
                        "", 
                        "", 
                        "无",
                        "无",
                        "无",
                        0, 
                        f"{crawl_duration} 秒", 
                        f"错误: {error_msg}",
                        retries
                    )
                    error_count += 1
        except requests.exceptions.Timeout:
            print(f"索引 {index} 爬取超时")
            retries += 1
            if retries > max_retries:
                save_result_to_excel(
                    index, 
                    original_links.get(index, "未知链接"),
                    "",
                    "", 
                    "超时",
                    "无",
                    "无",
                    0, 
                    "", 
                    "错误: 请求超时",
                    retries
                )
                error_count += 1
        except requests.exceptions.RequestException as e:
            print(f"索引 {index} 请求异常: {str(e)}")
            retries += 1
            if retries > max_retries:
                save_result_to_excel(
                    index, 
                    original_links.get(index, "未知链接"),
                    "",
                    "", 
                    "请求异常",
                    "无",
                    "无",
                    0, 
                    "", 
                    f"错误: 请求异常 - {str(e)}",
                    retries
                )
                error_count += 1
        except Exception as e:
            print(f"索引 {index} 未知错误: {str(e)}")
            print(traceback.format_exc())
            retries += 1
            if retries > max_retries:
                save_result_to_excel(
                    index, 
                    original_links.get(index, "未知链接"),
                    "", 
                    "", 
                    "未知错误",
                    "无",
                    "无",
                    0, 
                    "", 
                    f"错误: 未知异常 - {str(e)}",
                    retries
                )
                error_count += 1
        finally:
            if index % save_interval == 0:
                wb.save(result_file)
                print(f"定期保存结果文件：{result_file}")
            print(f"完成索引 {index} 处理")
            print("-" * 40)

# 添加汇总行
if SAVE_TO_EXCEL:
    sheet.cell(row=row_counter, column=1).value = "汇总"
    sheet.cell(row=row_counter, column=4).value = f"总计爬取: {len(indices)}个链接, 成功: {success_count}, 失败: {error_count}, 重试: {retry_count}"
    wb.save(result_file)
    print(f"已完成所有爬取任务，结果保存在 {result_file}")
else:
    print(f"爬取任务完成，Excel输出功能已禁用")

# 另存一份原始脚本的Excel
try:
    if os.path.exists(result_file) and SAVE_TO_EXCEL:
        # 另存为"测试爬取全部数据.xlsx"
        backup_file = "测试爬取全部数据.xlsx"
        wb.save(backup_file)
        print(f"已另存一份Excel到 {backup_file}")
except Exception as e:
    print(f"保存备份文件时出错: {str(e)}")

print(f"爬取任务统计: 总计尝试爬取 {len(indices)} 个链接, 成功: {success_count}, 失败: {error_count}") 