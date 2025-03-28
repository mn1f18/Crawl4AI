import requests
import openpyxl
from datetime import datetime
import os
import sys
import asyncio
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig
from crawl4ai.content_filter_strategy import PruningContentFilter
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
from bs4 import BeautifulSoup
from urllib.parse import urlparse

SAVE_TO_EXCEL = True  # 设置为False可以禁用Excel输出功能



def create_result_excel():
    if not SAVE_TO_EXCEL:
        return None, None, None
        
    # 使用时间戳命名文件，避免覆盖
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = f'top20结果.xlsx'
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Crawl Results'
    
    # 设置标题行
    headers = ['索引', '名称', '链接', '次数','内部网址']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header
    
    wb.save(result_file)
    print(f"创建结果文件：{result_file}")
    return wb, sheet, result_file

# 创建Excel结果文件
wb, sheet, result_file = create_result_excel()

# 计数器，用于记录当前行
row_counter = 2  # 从第2行开始，第1行是标题
def create_impot_web(sheet):
    try:
            source_wb = openpyxl.load_workbook('Top20网址.xlsx')#重要数据网址
            source_sheet = source_wb.active
            indices = range(20) #设置重要网址的数量
            for i in indices:
                sheet.cell(row=i+2, column=1).value = source_sheet.cell(row=i+1, column=1).value
                sheet.cell(row=i+2, column=2).value = source_sheet.cell(row=i+1, column=2).value
                sheet.cell(row=i+2, column=3).value = source_sheet.cell(row=i+1, column=3).value  # Excel行从1开始，跳过标题行
                sheet.cell(row=i+2, column=4).value = 0


                    
    except Exception as e:
        print(f"读取原始链接出错: {str(e)}")
    return sheet
sheet = create_impot_web(sheet)
wb.save(result_file)

#确认域名相同
def is_same_domain(url1, url2):
    domain1 = urlparse(url1).netloc
    domain2 = urlparse(url2).netloc
    return domain1 == domain2

#先排列网页再查找网站
def check_number_of_import_web(sheet):
    try:
            source_wb = openpyxl.load_workbook('export (13).xlsx')
            source_sheet = source_wb.active
            rows = list(source_sheet.iter_rows(values_only=True))  

     
            headers = rows[0]
            data = rows[1:]


            sorted_data = sorted(data, key=lambda row: row[1]) 


            for col_num, header in enumerate(headers, 1):
                source_sheet.cell(row=1, column=col_num).value = header

            for row_num, row in enumerate(sorted_data, start=2):
                for col_num, value in enumerate(row, start=1):
                    source_sheet.cell(row=row_num, column=col_num).value = value
                          
    except Exception as e:
        print(f"读取原始链接出错: {str(e)}")
    
    #上方排列下方找到相同的网页
    #['索引', '名称', '链接', '次数', '内部网址']
    for row_num in range(sheet.max_row-1):
        for i in range(source_sheet.max_row):

            if is_same_domain(sheet.cell(row=row_num+2, column=3).value
                            , source_sheet.cell(row=i+1, column=4).value):
                if sheet.cell(row=row_num+2, column=5).value== None:
                    sheet.cell(row=row_num+2, column=5).value = source_sheet.cell(row=i+1, column=4).value
                else:    
                    sheet.cell(row=row_num+2, column=5).value= sheet.cell(row=row_num+2, column=5).value+","+(source_sheet.cell(row=i+1, column=4).value)
                    sheet.cell(row=row_num+2, column=4).value =int(sheet.cell(row=row_num+2, column=4).value)+1

              
         

    return sheet
sheet = check_number_of_import_web(sheet)
wb.save(result_file)
#整理成testsample.xlsx
def swap_to_test(sheet):
    try:
        source_wb = openpyxl.Workbook()
        source_sheet = source_wb.active
        source_wb.title='Crawl Results'
        source_wb.save('testsample.xlsx')
        # 设置标题行
        numberofweb = 2
        headers = ['Testlink']
        for col, header in enumerate(headers, 1):
            source_sheet.cell(row=1, column=col).value = header
        for i in range(sheet.max_row):
            if sheet.cell(row=i+2, column=5).value!=None:
                items = sheet.cell(row=i+2, column=5).value.split(',')
                for x, item in enumerate(items, start=1):

                    source_sheet.cell(row=numberofweb, column=1).value = item
                    print(source_sheet.cell(row=numberofweb, column=1).value)
                    numberofweb+=1
        source_wb.save('testsample.xlsx')
    except Exception as e:
        print(f"读取原始链接出错: {str(e)}")
    return sheet
sheet = swap_to_test(sheet)